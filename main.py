import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import asksaveasfilename
import os
import sys
import json
import logging
from typing import Optional, Tuple, List, Dict, Any
from dataclasses import dataclass
from datetime import datetime
from tkinter import font
import matplotlib.pyplot as plt
import ttkbootstrap as tb
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
import random
import pandas as pd
from datetime import timedelta

logger = logging.getLogger("SeedRateApp")
logger.setLevel(logging.DEBUG)
ch = logging.StreamHandler()
ch.setLevel(logging.DEBUG)
formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
ch.setFormatter(formatter)
logger.addHandler(ch)

CONFIG_FILENAME = os.path.join(os.path.expanduser("~"), ".seed_rate_app_config.json")

CM_TO_M_FACTOR = 100
KG_PER_TON = 1000


@dataclass
class SeedRateParameters:
    wheel_diameter: float
    working_width: float
    total_seeds_in_20_rotations: float

    def seed_rate_per_decare(self) -> float:
        denominator = 0.063 * self.wheel_diameter * self.working_width
        if denominator == 0:
            logger.warning("Denominator for seed rate calculation is zero.")
            return 0.0
        result = self.total_seeds_in_20_rotations / denominator
        logger.debug(f"Calculated seed rate per decare: {result}")
        return result


def calculate_field_seed_rate(q: float, L: float, B: float) -> float:
    if L == 0 or B == 0:
        logger.warning("L or B is zero in field seed rate calculation.")
        return 0.0
    result = (1000 * q) / (2 * L * B)
    logger.debug(f"Calculated field seed rate: {result}")
    return result


def calculate_marker_lengths(b: float, L: float, unit: str = 'cm') -> Tuple[float, float]:
    b_cm = b * CM_TO_M_FACTOR if unit == 'm' else b
    L_cm = L * CM_TO_M_FACTOR if unit == 'm' else L

    x_sag = b_cm - (L_cm / 2)
    x_sol = b_cm + (L_cm / 2)
    logger.debug(
        f"Marker lengths calculated from b={b}{unit}, L={L}{unit}: Right={x_sag} cm, Left={x_sol} cm")
    return x_sag, x_sol


def calculate_q20_from_params(Q: float, D: float, B: float) -> float:
    result = Q * 0.063 * D * B
    logger.debug(f"Calculated Q20 from Q={Q}, D={D}, B={B}: {result}")
    return result


def calculate_row_spacing(D: float, i: float, n: float) -> float:
    if i == 0 or n == 0:
        logger.warning("Transmisyon oranı veya delik sayısı sıfır olamaz.")
        return 0.0
    result = (3.14 * D) / (i * n)
    logger.debug(f"Calculated row spacing: {result}")
    return result


def calculate_fertilization_qh(vb: float, h: float, b_case: float, lambda_d: float) -> float:
    result = 0.06 * vb * h * b_case * lambda_d
    logger.debug(f"Calculated Q(h): {result}")
    return result


def calculate_fertilization_q(qh: float, W: float, V: float) -> float:
    if W == 0 or V == 0:
        logger.warning("İş genişliği veya ilerleme hızı sıfır olamaz.")
        return 0.0
    result = qh / (W * V)
    logger.debug(f"Calculated fertilization norm (ton/da): {result}")
    return result



def calculate_work_performance(work_width: float, speed: float, efficiency: float) -> float:
    """ Saatte etkili olarak işlenen alanı dekar cinsinden hesaplar. """
    if speed == 0 or work_width == 0:
        return 0.0
    # Formül: (İş Genişliği (m) * Hız (km/h) * Verimlilik (%)) / 1000
    # Sadeleştirilmiş hali: İş Genişliği * Hız * Verimlilik / 10
    result = (work_width * speed * (efficiency / 100)) * 1000 / 10000
    result = (work_width * speed * efficiency) / 10.0
    logger.debug(f"Calculated work performance: {result} da/saat")
    return result


def calculate_germination_seed_rate(target_plants: float, thousand_grain_weight: float, germination_rate: float,
                                    purity_rate: float) -> float:
    """ Çimlenme ve saflık oranına göre dekara atılması gereken tohum miktarını kg cinsinden hesaplar. """
    denominator = germination_rate * purity_rate
    if denominator == 0:
        logger.warning("Germination or purity rate is zero.")
        return 0.0
    # Formül: (Hedef Bitki Sayısı * 1000 Tane Ağırlığı) / (Çimlenme Oranı * Saflık Oranı * 100)
    # (gram/da cinsinden sonucu kg/da'ya çevirmek için 1000'e bölmek gerekir)
    result_gram_da = (target_plants * thousand_grain_weight) / ((germination_rate / 100) * (purity_rate / 100))
    result_kg_da = result_gram_da / 1000
    logger.debug(f"Calculated seed rate based on germination: {result_kg_da} kg/da")
    return result_kg_da


def parse_float(value: str) -> Optional[float]:
    if not isinstance(value, str):
        logger.debug(f"Non-string value received in parse_float: {value}")
        return None
    cleaned_value = value.strip().replace(",", ".")
    if not cleaned_value:
        logger.debug("Empty string after cleaning in parse_float.")
        return None
    try:
        return float(cleaned_value)
    except ValueError as e:
        logger.error(f"Failed to parse float from value: '{value}'. Error: {e}")
        return None


def is_float_input_valid(value: str) -> bool:
    if not isinstance(value, str):
        return False
    cleaned_value = value.strip().replace(",", ".")
    if not cleaned_value:
        return True
    try:
        float(cleaned_value)
        return True
    except ValueError:
        return False


@dataclass
class CalculationResult:
    calculation_type: str
    inputs: Dict[str, str]
    results: Dict[str, str]
    timestamp: str


class Settings:
    def __init__(self, config_path: str = CONFIG_FILENAME):
        self.config_path = config_path
        self.data = {"unit": "m", "history": {}}
        self.load()

    def load(self) -> None:
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, "r", encoding="utf-8") as f:
                    self.data = json.load(f)
                logger.info("Settings loaded from config file.")
                if "history" not in self.data:
                    self.data["history"] = {}
                for tab_name, entries in self.data["history"].items():
                    new_entries = []
                    for entry in entries:
                        if isinstance(entry, dict) and "timestamp" not in entry:
                            if not isinstance(entry, dict):
                                new_entries.append(CalculationResult(
                                    calculation_type=tab_name,
                                    inputs={"old_entry_data": entry},
                                    results={"old_result_data": entry},
                                    timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                                ).__dict__)
                            else:
                                new_entries.append(entry)
                        else:
                            new_entries.append(entry)
                    self.data["history"][tab_name] = new_entries
            except Exception as e:
                logger.error(f"Error loading config file: {e}")
                self.data = {"unit": "m", "history": {}}

    def save(self) -> None:
        try:
            with open(self.config_path, "w", encoding="utf-8") as f:
                json.dump(self.data, f, ensure_ascii=False, indent=4)
            logger.info("Settings saved to config file.")
        except Exception as e:
            logger.error(f"Error saving config file: {e}")

    def get(self, key: str, default=None):
        return self.data.get(key, default)

    def set(self, key: str, value) -> None:
        self.data[key] = value
        self.save()

    def add_history(self, tab_name: str, calculation_result: CalculationResult) -> None:
        if tab_name not in self.data["history"]:
            self.data["history"][tab_name] = []
        self.data["history"][tab_name].insert(0, calculation_result.__dict__)
        self.data["history"][tab_name] = self.data["history"][tab_name][:5]
        self.save()

    def get_history(self, tab_name: str) -> List[CalculationResult]:
        history_list_of_dicts = self.data["history"].get(tab_name, [])
        return [CalculationResult(**res_dict) for res_dict in history_list_of_dicts]


class PDFGenerator:
    FONT_NAME_FOR_PDF = "DejaVuSans"

    @staticmethod
    def generate_pdf(filename: str, data: dict) -> bool:
        try:
            font_path = os.path.join(os.path.dirname(__file__), "DejaVuSans.ttf")
            if not os.path.exists(font_path):
                logger.error(f"DejaVuSans.ttf not found at {font_path}. Using default font.")
                PDFGenerator.FONT_NAME_FOR_PDF = "Helvetica"
            else:
                try:
                    pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))
                    PDFGenerator.FONT_NAME_FOR_PDF = "DejaVuSans"
                    logger.info("DejaVuSans font registered successfully for PDF.")
                except Exception as e:
                    logger.error(f"Failed to register DejaVuSans font for PDF: {e}. Using default font.")
                    PDFGenerator.FONT_NAME_FOR_PDF = "Helvetica"

            doc = SimpleDocTemplate(filename, pagesize=A4)
            styles = getSampleStyleSheet()

            for style_name in ['Normal', 'BodyText', 'Heading1', 'Heading2', 'Title', 'Italic']:
                if style_name in styles:
                    styles[style_name].fontName = PDFGenerator.FONT_NAME_FOR_PDF

            table_style = TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), PDFGenerator.FONT_NAME_FOR_PDF),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B444B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (1, -1), 'CENTER'),
                ('ALIGN', (2, 0), (2, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), PDFGenerator.FONT_NAME_FOR_PDF),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F5F5DC')),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6)
            ])

            story = []
            title_style = styles['Title']
            title_style.alignment = 1
            title = Paragraph(f"<b>{data['title']}</b>", title_style)
            story.append(title)
            story.append(Spacer(1, 24))

            if data.get('parameters'):
                param_data = [["Parametre", "Değer", "Birim"]]
                for param in data['parameters']:
                    param_data.append(param)
                param_table = Table(param_data, colWidths=[200, 150, 100])
                param_table.setStyle(table_style)
                story.append(Paragraph("<b>Giriş Parametreleri:</b>", styles['Normal']))
                story.append(Spacer(1, 6))
                story.append(param_table)
                story.append(Spacer(1, 24))

            if data.get('results'):
                story.append(Paragraph("<b>Hesaplama Sonuçları:</b>", styles['Normal']))
                story.append(Spacer(1, 6))
                for result in data['results']:
                    if isinstance(result, list):
                        story.append(Paragraph(f"<b>{result[0]}:</b> {result[1]}", styles['BodyText']))
                    else:
                        story.append(Paragraph(result, styles['BodyText']))
                    story.append(Spacer(1, 8))
                story.append(Spacer(1, 24))

            if 'formula' in data and data['formula']:
                story.append(Paragraph("<b>Kullanılan Formül:</b>", styles['Normal']))
                story.append(Spacer(1, 6))
                story.append(Paragraph(f"<font color='#555555'><i>{data['formula']}</i></font>", styles['Italic']))
                story.append(Spacer(1, 12))

            doc.build(story)
            return True
        except Exception as e:
            import traceback
            logger.error(f"PDF generation error: {e}\n{traceback.format_exc()}")
            return False


class SeedRateApp(tb.Window):
    def __init__(self):
        super().__init__(themename="superhero")
        self.title("Ekim ve Gübreleme Makineleri Kalibrasyon Aracı")

        try:
            base_path = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
            icon_path = os.path.join(base_path, "appicon.ico")
            if not os.path.exists(icon_path):
                print(f"Icon file expected at: {icon_path} but not found.")
            if os.path.exists(icon_path):
                self.iconbitmap(icon_path)
            else:
                logger.warning(f"Icon file not found at: {icon_path}")
        except Exception as e:
            logger.warning(f"Icon not set due to error: {e}")

        self.geometry("700x1000")
        self.resizable(False, False)

        self.settings = Settings()
        self.unit = self.settings.get("unit", "m")
        self.unit_var = tb.StringVar(value=self.unit)
        self.labels_to_update = []
        self.history_windows: Dict[str, Optional[tb.Toplevel]] = {}

        try:
            default_font = font.nametofont("TkDefaultFont")
            default_font.configure(family="DejaVu Sans", size=10)
            text_font = font.nametofont("TkTextFont")
            text_font.configure(family="DejaVu Sans", size=10)
            menu_font = font.nametofont("TkMenuFont")
            menu_font.configure(family="DejaVu Sans", size=10)
            heading_font = font.nametofont("TkHeadingFont")
            heading_font.configure(family="DejaVu Sans", size=11, weight="bold")

            style = tb.Style()
            style.configure('.', font=("DejaVu Sans", 10))
            style.configure('TButton', font=("DejaVu Sans", 10, "bold"))
            style.configure('TLabel', font=("DejaVu Sans", 10))
            style.configure('TLabelframe.Label', font=("DejaVu Sans", 11, "bold"))
            style.configure('TEntry', font=("DejaVu Sans", 10))
            style.configure('TCombobox', font=("DejaVu Sans", 10))
            style.configure('TCheckbutton', font=("DejaVu Sans", 10))
            style.configure('TRadiobutton', font=("DejaVu Sans", 10))
            style.configure('TNotebook.Tab', font=("DejaVu Sans", 10, "bold"))
            logger.info("DejaVu Sans font configured for Tkinter/ttkbootstrap.")
        except Exception as e:
            logger.error(
                f"Failed to configure DejaVu Sans font for Tkinter/ttkbootstrap: {e}. Using default system font.")

        self.entry_widgets: Dict[str, List[tk.Widget]] = {
            "Ekim Normu": [],
            "Tarlada Ekim Normu": [],
            "Markör Uzunluğu": [],
            "Q20 Hesaplama": [],
            "Gübreleme Normu": [],
            "Sıra Üzeri Uzaklık": [],
            "Ekim Kalitesi": [],
            "İş Başarısı": [],
            "Çimlenme Normu": []
        }

        self.last_calculation_results: Dict[str, Any] = {}
        self.create_widgets()

    def create_widgets(self):
        self.create_quick_access_menu()
        top_buttons_frame = tb.Frame(self)
        top_buttons_frame.pack(side="top", pady=10)

        help_button = tb.Button(top_buttons_frame, text="Yardım",
                                command=self.show_help_window, bootstyle="secondary")
        help_button.pack(side="left", padx=5)

        stats_button = tb.Button(top_buttons_frame, text="📊 İstatistikler",
                                 command=self.show_quick_stats, bootstyle="secondary")
        stats_button.pack(side="left", padx=5)

        save_all_button = tb.Button(top_buttons_frame, text="   Tümünü Kaydet",
                                    command=self.save_all_calculations, bootstyle="secondary")
        save_all_button.pack(side="left", padx=5)

        clear_history_button = tb.Button(top_buttons_frame, text="🗑️ Geçmişi Temizle",
                                         command=self.clear_all_history, bootstyle="secondary")
        clear_history_button.pack(side="left", padx=5)

        self.control_frame = tb.Frame(self)
        self.control_frame.pack(side="top", fill="x", pady=5, padx=10)
        self.control_frame.pack_forget()

        self.unit_frame = tb.Frame(self.control_frame)
        self.unit_frame.pack(side="left")

        self.back_btn_frame = tb.Frame(self.control_frame)
        self.back_btn_frame.pack(side="right")

        self.notebook_frame = tb.Frame(self)
        self.notebook_frame.pack(fill='both', expand=True, padx=10, pady=(10, 0))

        self.show_selection_screen()

    def create_unit_switch(self):
        for widget in self.unit_frame.winfo_children():
            widget.destroy()
        tb.Label(self.unit_frame, text="Birim Seçimi:").pack(side="left")
        unit_menu = tb.OptionMenu(
            self.unit_frame, self.unit_var, self.unit, "cm", "m", command=self.update_units)
        unit_menu.pack(side="left", padx=5)

    def update_units(self, value: str):
        old_unit = self.unit
        self.unit = value
        self.settings.set("unit", value)
        unit_text = f"({self.unit})"
        valid_labels = []
        for label in self.labels_to_update:
            try:
                if label.winfo_exists():
                    valid_labels.append(label)
                    text = label.cget("text")
                    base_text = text.split('(')[0].strip()
                    new_text = f"{base_text} {unit_text}"
                    label.config(text=new_text)
            except tk.TclError:
                continue
        self.labels_to_update = valid_labels
        logger.debug(f"Units updated from {old_unit} to {self.unit}")

    def show_tabs(self, option: str):
        for widget in self.notebook_frame.winfo_children():
            widget.destroy()

        self.control_frame.pack(side="top", fill="x", pady=5, padx=10)
        self.create_unit_switch()

        for widget in self.back_btn_frame.winfo_children():
            widget.destroy()
        back_btn = tb.Button(
            self.back_btn_frame, text="Seçim Ekranına Geri Dön",
            command=self.show_selection_screen, bootstyle="info")
        back_btn.pack()

        self.notebook = tb.Notebook(self.notebook_frame)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=(10, 0))

        self.entry_widgets = {
            "Ekim Normu": [], "Tarlada Ekim Normu": [], "Markör Uzunluğu": [],
            "Q20 Hesaplama": [], "Gübreleme Normu": [], "Sıra Üzeri Uzaklık": [],
            "Ekim Kalitesi": [], "İş Başarısı": [], "Çimlenme Normu": []
        }

        if option == "tahil":
            self.create_tab4_q20()
            self.create_tab1()
            self.create_tab2()
            self.create_tab3()
        elif option == "hassas":
            self.create_sira_uzeri_tab()
            self.create_tab_ekim_kalitesi()
            self.create_tab3()
        elif option == "Gübre":
            self.create_tab_fertilization()
        elif option == "Diğer":
            self.create_other_calculations_tab()

        self.update_units(self.unit)

    def show_selection_screen(self):
        self.control_frame.pack_forget()
        for widget in self.notebook_frame.winfo_children():
            widget.destroy()

        tb.Label(self.notebook_frame, text="Lütfen Bir Makine Tipi Seçiniz:",
                 font=("DejaVu Sans", 16, "bold")).pack(pady=(50, 20))

        btn_width = 30
        ipad_x = 40
        ipad_y = 20
        pady_val = 15

        tahil_btn = tb.Button(
            self.notebook_frame, text="Tahıl Ekim Makineleri", width=btn_width,
            command=lambda: self.show_tabs("tahil"), bootstyle="primary")
        tahil_btn.pack(pady=pady_val, ipadx=ipad_x, ipady=ipad_y)

        hassas_btn = tb.Button(
            self.notebook_frame, text="Hassas Ekim Makineleri", width=btn_width,
            command=lambda: self.show_tabs("hassas"), bootstyle="primary")
        hassas_btn.pack(pady=pady_val, ipadx=ipad_x, ipady=ipad_y)

        gubre_btn = tb.Button(
            self.notebook_frame, text="Gübre Dağıtma Makineleri", width=btn_width,
            command=lambda: self.show_tabs("Gübre"), bootstyle="primary")
        gubre_btn.pack(pady=pady_val, ipadx=ipad_x, ipady=ipad_y)


        credits_frame = tb.Frame(self.notebook_frame)
        credits_frame.pack(side="bottom", pady=(20, 20))
        tb.Label(credits_frame, text="Hazırlayan: Hasan Dural",
                 font=("DejaVu Sans", 10, "italic"), foreground="gray").pack()
        tb.Label(credits_frame, text="Danışman: Prof.Dr.Davut Karayel",
                 font=("DejaVu Sans", 10, "italic"), foreground="gray").pack()

    def show_help_window(self):
        if hasattr(self, 'help_window') and self.help_window.winfo_exists():
            self.help_window.lift()
            return

        self.help_window = tb.Toplevel(self)
        self.help_window.title("Yardım")
        self.help_window.geometry("550x650")
        self.help_window.resizable(False, False)

        help_text = (
            "<b>Ekim ve Gübreleme Hesaplamaları Hakkında Bilgiler:</b>\n\n"
            "<b>1. Ekim Normu Hesaplama (20 Devirdeki Tohum Miktarı Biliniyorsa):</b>\n"
            "   Formül: Q20=Q×0.063×D×B\n"
            "   - **Q20:** 20 tekerlek devrinde atılan tohum miktarı (kg)\n"
            "   - **D:** Tekerlek çapı (m)\n"
            "   - **B:** İş genişliği (m)\n"
            "   - **Q:** Ekim Normu (kg/da)\n\n"
            "<b>2. Tarlada Yapılan Ekim Normu:</b>\n"
            "   Formül:  Q=(1000×q)/(2×L×B)\n"
            "   - **q:** Belirli bir alanda tüketilen tohum miktarı (kg)\n"
            "   - **L:** Ekim makinesinin aldığı yol (m)\n"
            "   - **B:** Ekim makinesinin iş genişliği (m)\n"
            "   - **Q:** Ekim Normu (kg/da)\n\n"
            "<b>3. Markör Uzunluğu Hesaplamaları:</b>\n"
            "   Formüller:\n"
            "   - **Sağ Markör (cm):= B−(L/2)\n"
            "   - **Sol Markör (cm):**= B+(L/2)\n"
            "   - **b:** Ekim makinesi iş genişliği (cm veya m, otomatik çevrilir)\n"
            "   - **L:** Traktör ön tekerlek iz genişliği (cm veya m, otomatik çevrilir)\n\n"
            "<b>4. Q20 Hesaplama (Ekim Normu Biliniyorsa - 20 Devirdeki Tohum Miktarı):</b>\n"
            "   Formül: Q20=Q×0.063×D×B\n"
            "   - **Q:** Ekim Normu (kg/da)\n"
            "   - **D:** Tekerlek çapı (m)\n"
            "   - **B:** İş genişliği (m)\n\n"
            "<b>5. Sıra Üzeri Uzaklık (Hassas Tarım Makineleri):</b>\n"
            "   Formül: a=(π×D)/(i×n\n"
            "   - **D:** Tekerlek çapı (m)\n"
            "   - **i:** Transmisyon oranı (boyutsuz)\n"
            "   - **n:** Delik sayısı (adet)\n\n"
            "<b>6. Ekim Kalitesi (Hassas Tarım Makineleri):</b>\n"
            "   Tarlada yapılan sıra üzeri mesafe ölçümlerine göre ekimin kalitesini değerlendirir.\n"
            "   - **Ayarlanan Mesafe: Makinede hedeflenen sıra üzeri mesafe.\n"
            "   - **İkizleme: Ölçülen mesafenin ayarlanan mesafenin yarısından az olması.\n"
            "   - **Boşluk:** Ölçülen mesafenin ayarlanan mesafenin bir buçuk katından fazla olması .\n"
            "   - **Kabul Edilebilir:** İkizleme veya boşluk olmayan tüm ölçümler.\n"
            "   - Belirlenen bir yüzdesel sınırın (örn: Mısır için %90, Sebze için %85) üzerinde kabul edilebilir ekim yüzdesi olması durumunda ekim 'UYGUN' kabul edilir.\n\n"
            "<b>7. Gübreleme Normu Hesaplamaları:</b>\n"
            "   Formüller:\n"
            "   - **Q(h) (ton/saat):Q(h)=0.06×Vb×h×b×λ\n"
            "   - **q (kg/da):q=(Q(h)/(W×V))×1000 n"
            "   - **Vb:** Besleme hızı (m/dak)\n"
            "   - **h:** Kasa doldurma yüksekliği (m)\n"
            "   - **b:** Kasa genişliği (m)\n"
            "   - **$\\lambda$:** Gübrenin özgül ağırlığı (kg/m³)\n"
            "   - **W:** İş genişliği (m)\n"
            "   - **V:** İlerleme hızı (km/saat)\n\n"
            "**Not:** Birimler 'cm' veya 'm' olarak ayarlanabilir. Markör uzunlukları her zaman 'cm' olarak gösterilir."
        )

        text_widget = tb.Text(self.help_window, wrap="word", height=15, width=50, font=("DejaVu Sans", 10))

        text_widget.tag_configure("bold", font=("DejaVu Sans", 10, "bold"))
        text_widget.tag_configure("formula", font=("DejaVu Sans", 10, "italic"), foreground="blue")
        text_widget.tag_configure("info_bold", font=("DejaVu Sans", 10, "bold"))

        lines = help_text.split('\n')
        for line in lines:
            if line.startswith("<b>") and line.endswith("</b>"):
                text_widget.insert(tk.END, line[3:-4] + '\n\n', "bold")
            elif line.startswith("   Formül: $") and line.endswith("$"):
                display_line = line.replace("$\\text{", "").replace("}$", "").replace("\\times", "x").replace("\\pi",
                                                                                                              "π").replace(
                    "\\lambda", "λ").replace(" / ", " / ").strip()
                text_widget.insert(tk.END, "   Formül: " + display_line + '\n', "formula")
            elif line.startswith("   - **"):
                parts = line.split(":**", 1)
                if len(parts) == 2:
                    text_widget.insert(tk.END, "   - ", "")
                    text_widget.insert(tk.END, parts[0][7:] + ":", "info_bold")
                    text_widget.insert(tk.END, parts[1] + '\n', "")
                else:
                    text_widget.insert(tk.END, line + '\n', "")
            else:
                text_widget.insert(tk.END, line + '\n', "")

        text_widget.configure(state="disabled")
        text_widget.pack(pady=10, padx=10, fill="both", expand=True)

    def create_tab1(self):
        tab_name = "Ekim Normu"
        tab = tb.Frame(self.notebook)
        self.notebook.add(tab, text=tab_name)

        vcmd_float = (self.register(is_float_input_valid), '%P')

        input_frame = tb.LabelFrame(tab, text="Giriş Değerleri")
        input_frame.pack(padx=20, pady=10, fill="x")

        self.lbl_d = tb.Label(input_frame, text=f"Tekerlek Çapı ({self.unit})")
        self.lbl_d.pack(pady=(5, 0))
        self.entry_d = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_d.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_d)

        self.lbl_b = tb.Label(input_frame, text=f"İş Genişliği ({self.unit})")
        self.lbl_b.pack(pady=(5, 0))
        self.entry_b = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_b.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_b)

        tb.Label(input_frame, text="20 Devirdeki Tohum Miktarı (kg)").pack(pady=(5, 0))
        self.entry_s = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_s.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_s)

        button_frame = tb.Frame(tab)
        button_frame.pack(pady=10)
        tb.Button(button_frame, text="Hesapla", command=lambda: self.calculate_q20_tab(tab_name),
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame, text="Temizle", command=lambda: self.clear_entries(tab_name),
                  bootstyle="secondary").pack(side="left", padx=5)
        tb.Button(button_frame, text="Geçmiş", command=lambda: self.show_history_window(tab_name),
                  bootstyle="info").pack(side="left", padx=5)

        result_frame = tb.LabelFrame(tab, text="Sonuçlar")
        result_frame.pack(padx=20, pady=10, fill="x")
        self.result_label1 = tb.Label(result_frame, text="", font=('DejaVu Sans', 12))
        self.result_label1.pack(pady=5)

        area_frame = tb.LabelFrame(tab, text="Tarla Alanına Göre Toplam Tohum Miktarı")
        area_frame.pack(padx=20, pady=10, fill="x")
        tb.Label(area_frame, text="Tarla Alanı (da):").pack(pady=(5, 0))
        self.area_entry1 = tb.Entry(area_frame, font=("DejaVu Sans", 14), validate="key", validatecommand=vcmd_float)
        self.area_entry1.pack(pady=2)
        self.entry_widgets[tab_name].append(self.area_entry1)

        tb.Button(area_frame, text="Toplam Tohumu Hesapla", command=lambda: self.calculate_total_q20(tab_name),
                  bootstyle="info").pack(pady=5)

        button_frame_pdf = tb.Frame(tab)
        button_frame_pdf.pack(pady=10)

        tb.Button(button_frame_pdf, text="PDF'e Yazdır", command=lambda: self.save_pdf(1, tab_name),
                  bootstyle="warning").pack(side="left", padx=5)
        tb.Button(button_frame_pdf, text="Excel'e Aktar",
                  command=lambda: self.export_current_calculation_to_excel(tab_name), bootstyle="success").pack(
            side="left", padx=5)

        formul_label_q20 = tb.Label(tab, text="Formül: Q (kg/da) = Q20 / (0.063 × D (m) × B (m))",
                                    font=("DejaVu Sans", 10, "italic"), foreground="green")
        formul_label_q20.pack(pady=(10, 10))

        self.labels_to_update.extend([self.lbl_d, self.lbl_b])

    def calculate_q20_tab(self, tab_name: str):
        self.result_label1.config(text="")
        try:
            D_str = self.entry_d.get()
            B_str = self.entry_b.get()
            S_str = self.entry_s.get()

            if not D_str or not B_str or not S_str:
                messagebox.showerror("Hata", "Tüm giriş alanları doldurulmalıdır.")
                return
            if not is_float_input_valid(D_str) or not is_float_input_valid(B_str) or not is_float_input_valid(S_str):
                messagebox.showerror("Hata", "Geçersiz sayısal giriş.")
                return

            D = parse_float(D_str)
            B = parse_float(B_str)
            S = parse_float(S_str)

            D_calc = D / CM_TO_M_FACTOR if self.unit == "cm" else D
            B_calc = B / CM_TO_M_FACTOR if self.unit == "cm" else B
            logger.debug(f"Converted units for Q20 calc: D={D_calc}m, B={B_calc}m")

            params = SeedRateParameters(wheel_diameter=D_calc, working_width=B_calc, total_seeds_in_20_rotations=S)
            self.q_result = params.seed_rate_per_decare()
            result_text = f"Ekim Normu: {self.q_result:.2f} kg/da"
            self.result_label1.config(text=result_text, bootstyle="success")
            logger.info(f"Calculated Q20 seed rate: {self.q_result:.2f} kg/da")

            inputs = {"Tekerlek Çapı": D_str + f" {self.unit}", "İş Genişliği": B_str + f" {self.unit}",
                      "20 Devirdeki Tohum Miktarı": S_str + " kg"}
            results = {"Ekim Normu": f"{self.q_result:.2f} kg/da"}
            self.last_calculation_results[tab_name] = {
                "title": f"{tab_name} Hesaplama Raporu",
                "parameters": [["Tekerlek Çapı", D_str, self.unit], ["İş Genişliği", B_str, self.unit],
                               ["20 Devirdeki Tohum Miktarı", S_str, "kg"]],
                "results": [["Ekim Normu", f"{self.q_result:.2f} kg/da"]],
                "formula": "Q (kg/da) = Q20 (kg) / (0.063 × D (m) × B (m))"
            }
            self.settings.add_history(tab_name, CalculationResult(tab_name, inputs, results,
                                                                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

        except (ValueError, TypeError) as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})")
            logger.error(f"Error in Q20 calculation: {e}")
        except Exception as e:
            messagebox.showerror("Hata", f"Beklenmeyen bir hata oluştu: {e}")
            logger.error(f"Unexpected error in Q20 calculation: {e}")

    def calculate_total_q20(self, tab_name: str):
        try:
            area_str = self.area_entry1.get()
            if not area_str:
                self.result_label1.config(text="Hata: Tarla alanını giriniz!", bootstyle="danger")
                return
            if not is_float_input_valid(area_str):
                self.result_label1.config(text="Hata: Geçersiz alan değeri!", bootstyle="danger")
                return
            area = parse_float(area_str)
            if area is None or area <= 0:
                self.result_label1.config(text="Hata: Alan değeri sıfırdan büyük olmalıdır!", bootstyle="danger")
                return
            if not hasattr(self, 'q_result') or self.q_result is None:
                self.result_label1.config(text="Hata: Önce ekim normunu hesaplayın!", bootstyle="danger")
                return
            total_seeds = self.q_result * area
            result_text = f"Ekim Normu: {self.q_result:.2f} kg/da\nToplam Tohum Miktarı: {total_seeds:.2f} kg"
            self.result_label1.config(text=result_text, bootstyle="success")
            self.settings.add_history(tab_name, CalculationResult(
                calculation_type="Toplam Tohum Hesaplama",
                inputs={"Tarla Alanı": area_str + " da", "Ekim Normu": f"{self.q_result:.2f} kg/da"},
                results={"Toplam Tohum Miktarı": f"{total_seeds:.2f} kg"},
                timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
            logger.info(f"Toplam tohum hesaplandı: {total_seeds:.2f} kg")
        except Exception as e:
            error_msg = f"Hesaplama hatası: {str(e)}"
            self.result_label1.config(text=error_msg, bootstyle="danger")
            logger.error(f"Error in total seed calculation: {e}")

    def create_tab2(self):
        tab_name = "Tarlada Ekim Normu"
        tab = tb.Frame(self.notebook)
        self.notebook.add(tab, text=tab_name)
        vcmd_float = (self.register(is_float_input_valid), '%P')
        input_frame = tb.LabelFrame(tab, text="Giriş Değerleri")
        input_frame.pack(padx=20, pady=10, fill="x")
        tb.Label(input_frame, text="Tüketilen Tohum Miktarı (kg)").pack(pady=(5, 0))
        self.entry_q_field = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_q_field.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_q_field)
        self.lbl_l_field = tb.Label(input_frame, text=f"Makinenin Aldığı Yol ({self.unit})")
        self.lbl_l_field.pack(pady=(5, 0))
        self.entry_l_field = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_l_field.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_l_field)
        self.lbl_b_field = tb.Label(input_frame, text=f"İş Genişliği ({self.unit})")
        self.lbl_b_field.pack(pady=(5, 0))
        self.entry_b_field = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_b_field.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_b_field)
        button_frame = tb.Frame(tab)
        button_frame.pack(pady=10)
        tb.Button(button_frame, text="Hesapla", command=lambda: self.calculate_field_seed_rate_tab(tab_name),
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame, text="Temizle", command=lambda: self.clear_entries(tab_name),
                  bootstyle="secondary").pack(side="left", padx=5)
        tb.Button(button_frame, text="Geçmiş", command=lambda: self.show_history_window(tab_name),
                  bootstyle="info").pack(side="left", padx=5)
        result_frame = tb.LabelFrame(tab, text="Sonuçlar")
        result_frame.pack(padx=20, pady=10, fill="x")
        self.result_label2 = tb.Label(result_frame, text="", font=('DejaVu Sans', 12))
        self.result_label2.pack(pady=5)
        area_frame = tb.LabelFrame(tab, text="Tarla Alanına Göre Toplam Tohum Miktarı")
        area_frame.pack(padx=20, pady=10, fill="x")
        tb.Label(area_frame, text="Tarla Alanı (da):").pack(pady=(5, 0))
        self.area_entry2 = tb.Entry(area_frame, font=("DejaVu Sans", 14), validate="key", validatecommand=vcmd_float)
        self.area_entry2.pack(pady=2)
        self.entry_widgets[tab_name].append(self.area_entry2)
        tb.Button(area_frame, text="Toplam Tohumu Hesapla", command=lambda: self.calculate_total_q_field(tab_name),
                  bootstyle="info").pack(pady=5)
        button_frame_pdf = tb.Frame(tab)
        button_frame_pdf.pack(pady=10)
        tb.Button(button_frame_pdf, text="PDF'e Yazdır", command=lambda: self.save_pdf(2, tab_name),
                  bootstyle="warning").pack(side="left", padx=5)
        tb.Button(button_frame_pdf, text="Excel'e Aktar",
                  command=lambda: self.export_current_calculation_to_excel(tab_name), bootstyle="success").pack(
            side="left", padx=5)
        formul_label_field = tb.Label(tab, text="Formül: Q (kg/da) = (1000 × q ) / (2 × L (m) × B (m))",
                                      font=("DejaVu Sans", 10, "italic"), foreground="green")
        formul_label_field.pack(pady=(10, 10))
        self.labels_to_update.extend([self.lbl_l_field, self.lbl_b_field])

    def calculate_total_q_field(self, tab_name: str):
        try:
            area_str = self.area_entry2.get()
            if not area_str:
                self.result_label2.config(text="Hata: Tarla alanını giriniz!", bootstyle="danger")
                return
            if not is_float_input_valid(area_str):
                self.result_label2.config(text="Hata: Geçersiz alan değeri!", bootstyle="danger")
                return
            area = parse_float(area_str)
            if area is None or area <= 0:
                self.result_label2.config(text="Hata: Alan değeri sıfırdan büyük olmalıdır!", bootstyle="danger")
                return
            if not hasattr(self, 'q_field_result') or self.q_field_result is None:
                self.result_label2.config(text="Hata: Önce ekim normunu hesaplayın!", bootstyle="danger")
                return
            total_seeds = self.q_field_result * area
            result_text = f"Ekim Normu: {self.q_field_result:.2f} kg/da\nToplam Tohum Miktarı: {total_seeds:.2f} kg"
            self.result_label2.config(text=result_text, bootstyle="success")
            logger.info(f"Toplam tohum hesaplandı: {total_seeds:.2f} kg")
        except Exception as e:
            error_msg = f"Hesaplama hatası: {str(e)}"
            self.result_label2.config(text=error_msg, bootstyle="danger")
            logger.error(f"Error in total seed calculation for field: {e}")

    def calculate_field_seed_rate_tab(self, tab_name: str):
        self.result_label2.config(text="")
        try:
            q_field_str = self.entry_q_field.get()
            l_field_str = self.entry_l_field.get()
            b_field_str = self.entry_b_field.get()
            if not q_field_str or not l_field_str or not b_field_str:
                messagebox.showerror("Hata", "Tüm giriş alanları doldurulmalıdır.")
                return
            if not is_float_input_valid(q_field_str) or not is_float_input_valid(
                    l_field_str) or not is_float_input_valid(b_field_str):
                messagebox.showerror("Hata", "Geçersiz sayısal giriş.")
                return
            q_field = parse_float(q_field_str)
            l_field = parse_float(l_field_str)
            b_field = parse_float(b_field_str)
            l_field_calc = l_field / CM_TO_M_FACTOR if self.unit == "cm" else l_field
            b_field_calc = b_field / CM_TO_M_FACTOR if self.unit == "cm" else b_field
            logger.debug(f"Converted units for field seed rate calc: L={l_field_calc}m, B={b_field_calc}m")
            self.q_field_result = calculate_field_seed_rate(q_field, l_field_calc, b_field_calc)
            result_text = f"Tarlada Ekim Normu: {self.q_field_result:.2f} kg/da"
            self.result_label2.config(text=result_text, bootstyle="success")
            logger.info(f"Calculated field seed rate: {self.q_field_result:.2f} kg/da")
            inputs = {"Tüketilen Tohum Miktarı": q_field_str + " kg",
                      "Makinenin Aldığı Yol": l_field_str + f" {self.unit}",
                      "İş Genişliği": b_field_str + f" {self.unit}"}
            results = {"Tarlada Ekim Normu": f"{self.q_field_result:.2f} kg/da"}
            self.last_calculation_results[tab_name] = {
                "title": f"{tab_name} Hesaplama Raporu",
                "parameters": [["Tüketilen Tohum Miktarı", q_field_str, "kg"],
                               ["Makinenin Aldığı Yol", l_field_str, self.unit],
                               ["İş Genişliği", b_field_str, self.unit]],
                "results": [["Tarlada Ekim Normu", f"{self.q_field_result:.2f} kg/da"]],
                "formula": "Q (kg/da) = (1000 × q (kg)) / (2 × L (m) × B (m))"
            }
            self.settings.add_history(tab_name, CalculationResult(tab_name, inputs, results,
                                                                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        except (ValueError, TypeError) as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})")
            logger.error(f"Error in field seed rate calculation: {e}")
        except Exception as e:
            messagebox.showerror("Hata", f"Beklenmeyen bir hata oluştu: {e}")
            logger.error(f"Unexpected error in field seed rate calculation: {e}")

    def create_tab3(self):
        tab_name = "Markör Uzunluğu"
        tab = tb.Frame(self.notebook)
        self.notebook.add(tab, text=tab_name)
        vcmd_float = (self.register(is_float_input_valid), '%P')
        input_frame = tb.LabelFrame(tab, text="Giriş Değerleri")
        input_frame.pack(padx=20, pady=10, fill="x")
        self.lbl_b_marker = tb.Label(input_frame, text=f"Ekim Makinesi İş Genişliği ({self.unit})")
        self.lbl_b_marker.pack(pady=(5, 0))
        self.entry_b_marker = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_b_marker.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_b_marker)
        self.lbl_l_marker = tb.Label(input_frame, text=f"Traktör Ön Tekerlek İz Genişliği ({self.unit})")
        self.lbl_l_marker.pack(pady=(5, 0))
        self.entry_l_marker = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_l_marker.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_l_marker)
        button_frame = tb.Frame(tab)
        button_frame.pack(pady=10)
        tb.Button(button_frame, text="Hesapla", command=lambda: self.calculate_marker_lengths_tab(tab_name),
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame, text="Temizle", command=lambda: self.clear_entries(tab_name),
                  bootstyle="secondary").pack(side="left", padx=5)
        tb.Button(button_frame, text="Geçmiş", command=lambda: self.show_history_window(tab_name),
                  bootstyle="info").pack(side="left", padx=5)
        result_frame = tb.LabelFrame(tab, text="Sonuçlar")
        result_frame.pack(padx=20, pady=10, fill="x")
        self.result_label3_right = tb.Label(result_frame, text="", font=('DejaVu Sans', 12))
        self.result_label3_right.pack(pady=5)
        self.result_label3_left = tb.Label(result_frame, text="", font=('DejaVu Sans', 12))
        self.result_label3_left.pack(pady=5)
        button_frame_pdf = tb.Frame(tab)
        button_frame_pdf.pack(pady=10)
        tb.Button(button_frame_pdf, text="PDF'e Yazdır", command=lambda: self.save_pdf(3, tab_name),
                  bootstyle="warning").pack(side="left", padx=5)
        tb.Button(button_frame_pdf, text="Excel'e Aktar",
                  command=lambda: self.export_current_calculation_to_excel(tab_name), bootstyle="success").pack(
            side="left", padx=5)
        formul_label_marker_right = tb.Label(tab, text="Formül (Sağ Markör): b (cm) - (L (cm) / 2)",
                                             font=("DejaVu Sans", 10, "italic"), foreground="green")
        formul_label_marker_right.pack(pady=(10, 0))
        formul_label_marker_left = tb.Label(tab, text="Formül (Sol Markör): b (cm) + (L (cm) / 2)",
                                            font=("DejaVu Sans", 10, "italic"), foreground="green")
        formul_label_marker_left.pack(pady=(0, 10))
        self.labels_to_update.extend([self.lbl_b_marker, self.lbl_l_marker])

    def calculate_marker_lengths_tab(self, tab_name: str):
        self.result_label3_right.config(text="")
        self.result_label3_left.config(text="")
        try:
            b_marker_str = self.entry_b_marker.get()
            l_marker_str = self.entry_l_marker.get()
            if not b_marker_str or not l_marker_str:
                messagebox.showerror("Hata", "Tüm giriş alanları doldurulmalıdır.")
                return
            if not is_float_input_valid(b_marker_str) or not is_float_input_valid(l_marker_str):
                messagebox.showerror("Hata", "Geçersiz sayısal giriş.")
                return
            b_marker = parse_float(b_marker_str)
            l_marker = parse_float(l_marker_str)
            right_marker, left_marker = calculate_marker_lengths(b_marker, l_marker, self.unit)
            self.result_label3_right.config(text=f"Sağ Markör: {right_marker:.2f} cm", bootstyle="success")
            self.result_label3_left.config(text=f"Sol Markör: {left_marker:.2f} cm", bootstyle="success")
            logger.info(f"Calculated marker lengths: Right={right_marker:.2f} cm, Left={left_marker:.2f} cm")
            inputs = {"Ekim Makinesi İş Genişliği": b_marker_str + f" {self.unit}",
                      "Traktör Ön Tekerlek İz Genişliği": l_marker_str + f" {self.unit}"}
            results = {"Sağ Markör": f"{right_marker:.2f} cm", "Sol Markör": f"{left_marker:.2f} cm"}
            self.last_calculation_results[tab_name] = {
                "title": f"{tab_name} Hesaplama Raporu",
                "parameters": [["Ekim Makinesi İş Genişliği", b_marker_str, self.unit],
                               ["Traktör Ön Tekerlek İz Genişliği", l_marker_str, self.unit]],
                "results": [["Sağ Markör", f"{right_marker:.2f} cm"], ["Sol Markör", f"{left_marker:.2f} cm"]],
                "formula": "Sağ Markör (cm) = b (cm) - (L (cm) / 2)\nSol Markör (cm) = b (cm) + (L (cm) / 2)"
            }
            self.settings.add_history(tab_name, CalculationResult(tab_name, inputs, results,
                                                                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        except (ValueError, TypeError) as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})")
            logger.error(f"Error in marker length calculation: {e}")
        except Exception as e:
            messagebox.showerror("Hata", f"Beklenmeyen bir hata oluştu: {e}")
            logger.error(f"Unexpected error in marker length calculation: {e}")

    def create_tab4_q20(self):
        tab_name = "Q20 Hesaplama"
        tab = tb.Frame(self.notebook)
        self.notebook.add(tab, text=tab_name)
        vcmd_float = (self.register(is_float_input_valid), '%P')
        input_frame = tb.LabelFrame(tab, text="Giriş Değerleri")
        input_frame.pack(padx=20, pady=10, fill="x")
        tb.Label(input_frame, text="Ekim Normu (kg/da)").pack(pady=(5, 0))
        self.entry_q_calc = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_q_calc.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_q_calc)
        self.lbl_d_calc = tb.Label(input_frame, text=f"Tekerlek Çapı ({self.unit})")
        self.lbl_d_calc.pack(pady=(5, 0))
        self.entry_d_calc = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_d_calc.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_d_calc)
        self.lbl_b_calc = tb.Label(input_frame, text=f"İş Genişliği ({self.unit})")
        self.lbl_b_calc.pack(pady=(5, 0))
        self.entry_b_calc = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_b_calc.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_b_calc)
        button_frame = tb.Frame(tab)
        button_frame.pack(pady=10)
        tb.Button(button_frame, text="Hesapla", command=lambda: self.calculate_q20_from_q_tab(tab_name),
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame, text="Temizle", command=lambda: self.clear_entries(tab_name),
                  bootstyle="secondary").pack(side="left", padx=5)
        tb.Button(button_frame, text="Geçmiş", command=lambda: self.show_history_window(tab_name),
                  bootstyle="info").pack(side="left", padx=5)
        result_frame = tb.LabelFrame(tab, text="Sonuçlar")
        result_frame.pack(padx=20, pady=10, fill="x")
        self.result_label4 = tb.Label(result_frame, text="", font=('DejaVu Sans', 12))
        self.result_label4.pack(pady=5)
        button_frame_pdf = tb.Frame(tab)
        button_frame_pdf.pack(pady=10)
        tb.Button(button_frame_pdf, text="PDF'e Yazdır", command=lambda: self.save_pdf(4, tab_name),
                  bootstyle="warning").pack(side="left", padx=5)
        tb.Button(button_frame_pdf, text="Excel'e Aktar",
                  command=lambda: self.export_current_calculation_to_excel(tab_name), bootstyle="success").pack(
            side="left", padx=5)
        formul_label_q20_calc = tb.Label(tab, text="Formül: Q20 (kg) = Q (kg/da) × 0.063 × D (m) × B (m)",
                                         font=("DejaVu Sans", 10, "italic"), foreground="green")
        formul_label_q20_calc.pack(pady=(10, 10))
        self.labels_to_update.extend([self.lbl_d_calc, self.lbl_b_calc])

    def calculate_q20_from_q_tab(self, tab_name: str):
        self.result_label4.config(text="")
        try:
            q_calc_str = self.entry_q_calc.get()
            d_calc_str = self.entry_d_calc.get()
            b_calc_str = self.entry_b_calc.get()
            if not q_calc_str or not d_calc_str or not b_calc_str:
                messagebox.showerror("Hata", "Tüm giriş alanları doldurulmalıdır.")
                return
            if not is_float_input_valid(q_calc_str) or not is_float_input_valid(d_calc_str) or not is_float_input_valid(
                    b_calc_str):
                messagebox.showerror("Hata", "Geçersiz sayısal giriş.")
                return
            q_calc = parse_float(q_calc_str)
            d_calc = parse_float(d_calc_str)
            b_calc = parse_float(b_calc_str)
            d_calc_m = d_calc / CM_TO_M_FACTOR if self.unit == "cm" else d_calc
            b_calc_m = b_calc / CM_TO_M_FACTOR if self.unit == "cm" else b_calc
            logger.debug(f"Converted units for Q20 from Q calc: D={d_calc_m}m, B={b_calc_m}m")
            self.q20_result_from_q = calculate_q20_from_params(q_calc, d_calc_m, b_calc_m)
            result_text = f"20 Devirdeki Tohum Miktarı (Q20): {self.q20_result_from_q:.2f} kg"
            self.result_label4.config(text=result_text, bootstyle="success")
            logger.info(f"Calculated Q20 from Q: {self.q20_result_from_q:.2f} kg")
            inputs = {"Ekim Normu": q_calc_str + " kg/da", "Tekerlek Çapı": d_calc_str + f" {self.unit}",
                      "İş Genişliği": b_calc_str + f" {self.unit}"}
            results = {"20 Devirdeki Tohum Miktarı (Q20)": f"{self.q20_result_from_q:.2f} kg"}
            self.last_calculation_results[tab_name] = {
                "title": f"{tab_name} Hesaplama Raporu",
                "parameters": [["Ekim Normu", q_calc_str, "kg/da"], ["Tekerlek Çapı", d_calc_str, self.unit],
                               ["İş Genişliği", b_calc_str, self.unit]],
                "results": [["20 Devirdeki Tohum Miktarı (Q20)", f"{self.q20_result_from_q:.2f} kg"]],
                "formula": "Q20 (kg) = Q (kg/da) × 0.063 × D (m) × B (m)"
            }
            self.settings.add_history(tab_name, CalculationResult(tab_name, inputs, results,
                                                                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        except (ValueError, TypeError) as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})")
            logger.error(f"Error in Q20 from Q calculation: {e}")
        except Exception as e:
            messagebox.showerror("Hata", f"Beklenmeyen bir hata oluştu: {e}")
            logger.error(f"Unexpected error in Q20 from Q calculation: {e}")

    def create_sira_uzeri_tab(self):
        tab_name = "Sıra Üzeri Uzaklık"
        tab = tb.Frame(self.notebook)
        self.notebook.add(tab, text=tab_name)
        vcmd_float = (self.register(is_float_input_valid), '%P')
        input_frame = tb.LabelFrame(tab, text="Giriş Değerleri")
        input_frame.pack(padx=20, pady=10, fill="x")
        self.lbl_d_row = tb.Label(input_frame, text=f"Tekerlek Çapı ({self.unit})")
        self.lbl_d_row.pack(pady=(5, 0))
        self.entry_d_row = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_d_row.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_d_row)
        tb.Label(input_frame, text="Transmisyon Oranı (i)").pack(pady=(5, 0))
        self.entry_i_row = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_i_row.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_i_row)
        tb.Label(input_frame, text="Delik Sayısı (n)").pack(pady=(5, 0))
        self.entry_n_row = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.entry_n_row.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_n_row)
        button_frame = tb.Frame(tab)
        button_frame.pack(pady=10)
        tb.Button(button_frame, text="Hesapla", command=lambda: self.calculate_row_spacing_tab(tab_name),
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame, text="Temizle", command=lambda: self.clear_entries(tab_name),
                  bootstyle="secondary").pack(side="left", padx=5)
        tb.Button(button_frame, text="Geçmiş", command=lambda: self.show_history_window(tab_name),
                  bootstyle="info").pack(side="left", padx=5)
        result_frame = tb.LabelFrame(tab, text="Sonuçlar")
        result_frame.pack(padx=20, pady=10, fill="x")
        self.result_label_row_spacing = tb.Label(result_frame, text="", font=('DejaVu Sans', 12))
        self.result_label_row_spacing.pack(pady=5)
        button_frame_pdf = tb.Frame(tab)
        button_frame_pdf.pack(pady=10)
        tb.Button(button_frame_pdf, text="PDF'e Yazdır", command=lambda: self.save_pdf(6, tab_name),
                  bootstyle="warning").pack(side="left", padx=5)
        tb.Button(button_frame_pdf, text="Excel'e Aktar",
                  command=lambda: self.export_current_calculation_to_excel(tab_name), bootstyle="success").pack(
            side="left", padx=5)
        formul_label_row_spacing = tb.Label(tab, text="Formül: a = (π × D (m)) / (i × n)",
                                            font=("DejaVu Sans", 10, "italic"), foreground="green")
        formul_label_row_spacing.pack(pady=(10, 10))
        self.labels_to_update.extend([self.lbl_d_row])

    def calculate_row_spacing_tab(self, tab_name: str):
        self.result_label_row_spacing.config(text="")
        try:
            d_row_str = self.entry_d_row.get()
            i_row_str = self.entry_i_row.get()
            n_row_str = self.entry_n_row.get()
            if not d_row_str or not i_row_str or not n_row_str:
                messagebox.showerror("Hata", "Tüm giriş alanları doldurulmalıdır.")
                return
            if not is_float_input_valid(d_row_str) or not is_float_input_valid(i_row_str) or not is_float_input_valid(
                    n_row_str):
                messagebox.showerror("Hata", "Geçersiz sayısal giriş.")
                return
            d_row = parse_float(d_row_str)
            i_row = parse_float(i_row_str)
            n_row = parse_float(n_row_str)
            d_row_m = d_row / CM_TO_M_FACTOR if self.unit == "cm" else d_row
            logger.debug(f"Converted units for row spacing calc: D={d_row_m}m")
            self.row_spacing_result = calculate_row_spacing(d_row_m, i_row, n_row)
            result_cm = self.row_spacing_result * CM_TO_M_FACTOR
            result_text = f"Sıra Üzeri Uzaklık: {self.row_spacing_result:.4f} m ({result_cm:.2f} cm)"
            self.result_label_row_spacing.config(text=result_text, bootstyle="success")
            logger.info(f"Calculated row spacing: {self.row_spacing_result:.4f} m ({result_cm:.2f} cm)")
            inputs = {"Tekerlek Çapı": d_row_str + f" {self.unit}", "Transmisyon Oranı": i_row_str,
                      "Delik Sayısı": n_row_str}
            results = {"Sıra Üzeri Uzaklık": f"{self.row_spacing_result:.4f} m ({result_cm:.2f} cm)"}
            self.last_calculation_results[tab_name] = {
                "title": f"{tab_name} Hesaplama Raporu",
                "parameters": [["Tekerlek Çapı", d_row_str, self.unit], ["Transmisyon Oranı", i_row_str, ""],
                               ["Delik Sayısı", n_row_str, "adet"]],
                "results": [["Sıra Üzeri Uzaklık", f"{self.row_spacing_result:.4f} m ({result_cm:.2f} cm)"]],
                "formula": "a (m) = (3.14 × D (m)) / (i × n)"
            }
            self.settings.add_history(tab_name, CalculationResult(tab_name, inputs, results,
                                                                  datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
        except (ValueError, TypeError) as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})")
            logger.error(f"Error in row spacing calculation: {e}")
        except Exception as e:
            messagebox.showerror("Hata", f"Beklenmeyen bir hata oluştu: {e}")
            logger.error(f"Unexpected error in row spacing calculation: {e}")

    def create_tab_ekim_kalitesi(self):
        tab_name = "Ekim Kalitesi"
        tab = tb.Frame(self.notebook)
        self.notebook.add(tab, text=tab_name)
        vcmd_float = (self.register(is_float_input_valid), '%P')
        input_frame = tb.LabelFrame(tab, text="Giriş Değerleri")
        input_frame.pack(padx=20, pady=10, fill="x")
        tb.Label(input_frame, text="Ayarlanan Sıra Üzeri Mesafe (cm)").pack(pady=(10, 2))
        self.ekim_ayarlanan_entry = tb.Entry(input_frame, validate="key", validatecommand=vcmd_float)
        self.ekim_ayarlanan_entry.pack()
        self.entry_widgets[tab_name].append(self.ekim_ayarlanan_entry)
        tb.Label(input_frame,
                 text="Tarlada Ölçülmüş Sıra Üzeri Mesafeler (cm)\n(Virgül ile ayırın veya her satıra bir değer yazın)").pack(
            pady=(10, 2))
        self.ekim_olcum_text = tk.Text(input_frame, height=5, width=40)
        self.ekim_olcum_text.pack()
        self.entry_widgets[tab_name].append(self.ekim_olcum_text)
        button_frame = tb.Frame(tab)
        button_frame.pack(pady=10)
        tb.Button(button_frame, text="Hesapla", command=lambda: self.ekim_kalitesi_hesapla(tab_name),
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame, text="Temizle", command=lambda: self.clear_entries(tab_name),
                  bootstyle="secondary").pack(side="left", padx=5)
        tb.Button(button_frame, text="Geçmiş", command=lambda: self.show_history_window(tab_name),
                  bootstyle="info").pack(side="left", padx=5)
        result_frame = tb.LabelFrame(tab, text="Sonuçlar")
        result_frame.pack(padx=20, pady=10, fill="x")
        self.ekim_kalitesi_sonuc = tb.Label(result_frame, text="", font=('Helvetica', 12), justify="left")
        self.ekim_kalitesi_sonuc.pack(pady=5)
        self.ekim_kalitesi_uygunluk_label = tb.Label(tab, text="", font=('Helvetica', 14, 'bold'))
        self.ekim_kalitesi_uygunluk_label.pack(pady=(10, 5))
        tb.Label(tab, text="Sınır Oranı Seçiniz:").pack(pady=(10, 2))
        self.ekim_sinir_var = tb.StringVar(value="90")
        tb.Radiobutton(tab, text="Mısır/Pamuk/Ayçiçeği (%90)", variable=self.ekim_sinir_var, value="90",
                       bootstyle="info").pack(anchor="w", padx=80)
        tb.Radiobutton(tab, text="Sebze (%85)", variable=self.ekim_sinir_var, value="85", bootstyle="info").pack(
            anchor="w", padx=80)
        button_frame_pdf = tb.Frame(tab)
        button_frame_pdf.pack(pady=10)
        tb.Button(button_frame_pdf, text="PDF'e Yazdır", command=lambda: self.save_pdf(7, tab_name),
                  bootstyle="warning").pack(side="left", padx=5)
        tb.Button(button_frame_pdf, text="Excel'e Aktar",
                  command=lambda: self.export_current_calculation_to_excel(tab_name), bootstyle="success").pack(
            side="left", padx=5)

    def ekim_kalitesi_hesapla(self, tab_name: str):
        self.ekim_kalitesi_sonuc.config(text="")
        self.ekim_kalitesi_uygunluk_label.config(text="", foreground="black")
        self.last_ekim_kalitesi_result = None
        try:
            ayarlanan_str = self.ekim_ayarlanan_entry.get()
            ayarlanan = parse_float(ayarlanan_str)
            if ayarlanan is None or ayarlanan <= 0:
                self.ekim_kalitesi_sonuc.config(
                    text="Hata: Lütfen geçerli bir Ayarlanan Sıra Üzeri Mesafe giriniz (> 0).", bootstyle="danger")
                logger.warning("Ekim Kalitesi: Geçersiz veya eksik ayarlanan mesafe girişi.")
                return
            raw_input_str = self.ekim_olcum_text.get("1.0", "end").strip()
            parts = raw_input_str.replace("\n", ",").split(",")
            olcumler = []
            for p in parts:
                val = parse_float(p.strip())
                if val is not None:
                    olcumler.append(val)
            if not olcumler:
                self.ekim_kalitesi_sonuc.config(text="Hata: Lütfen tarlada ölçülmüş sıra üzeri mesafeleri giriniz.",
                                                bootstyle="danger")
                logger.warning("Ekim Kalitesi: Ölçüm mesafeleri girilmedi.")
                return
            ikiz_sin = ayarlanan * 0.5
            bosluk_sin = ayarlanan * 1.5
            toplam = len(olcumler)
            ikiz = sum(1 for x in olcumler if x < ikiz_sin)
            bosluk = sum(1 for x in olcumler if x > bosluk_sin)
            kabul = toplam - ikiz - bosluk
            ikiz_yuzde = 100 * ikiz / toplam if toplam else 0
            bosluk_yuzde = 100 * bosluk / toplam if toplam else 0
            kabul_yuzde = 100 * kabul / toplam if toplam else 0
            secili_sinir_str = self.ekim_sinir_var.get()
            secili_sinir = parse_float(secili_sinir_str)
            if secili_sinir is None:
                self.ekim_kalitesi_sonuc.config(text="Hata: Lütfen bir sınır oranı seçiniz.", bootstyle="danger")
                logger.error("Ekim Kalitesi: Sınır oranı seçimi başarısız.")
                return
            sinir_metni = "Mısır/Pamuk/Ayçiçeği" if secili_sinir == 90 else "Sebze"
            sonuc = (
                f"Ayarlanan Sıra Üzeri Mesafe: {ayarlanan:.2f} cm\n"
                f"İkizleme Sınırı (< {ikiz_sin:.2f} cm)\n"
                f"Boşluk Sınırı (> {bosluk_sin:.2f} cm)\n"
                f"-----------------------------------------\n"
                f"Toplam Ölçüm: {toplam}\n"
                f"İkizleme Sayısı: {ikiz} (%{ikiz_yuzde:.1f})\n"
                f"Boşluk Sayısı: {bosluk} (%{bosluk_yuzde:.1f})\n"
                f"Kabul Edilebilir Sayısı: {kabul} (%{kabul_yuzde:.1f})\n"
                f"-----------------------------------------\n"
                f"Sınır Oranı ({sinir_metni}): %{secili_sinir:.0f}\n"
                f"Kabul Edilebilir Oran: %{kabul_yuzde:.1f}\n")
            self.ekim_kalitesi_sonuc.config(text=sonuc, bootstyle="info")
            uygun = kabul_yuzde >= secili_sinir
            if uygun:
                self.ekim_kalitesi_uygunluk_label.config(text="EKİM KALİTESİ: UYGUN ✅", foreground="green")
                logger.info(f"Ekim Kalitesi Uygun: {kabul_yuzde:.1f}% >= {secili_sinir}%")
            else:
                self.ekim_kalitesi_uygunluk_label.config(text="EKİM KALİTESİ: UYGUN DEĞİL ❌", foreground="red")
                logger.info(f"Ekim Kalitesi Uygun Değil: {kabul_yuzde:.1f}% < {secili_sinir}%")
            self.last_ekim_kalitesi_result = {"ayarlanan": ayarlanan, "ikiz_sin": ikiz_sin, "bosluk_sin": bosluk_sin,
                                              "toplam": toplam, "ikiz": ikiz, "ikiz_yuzde": ikiz_yuzde,
                                              "bosluk": bosluk, "bosluk_yuzde": bosluk_yuzde, "kabul": kabul,
                                              "kabul_yuzde": kabul_yuzde, "sinir": secili_sinir,
                                              "sinir_metni": sinir_metni, "uygun": uygun, "olcumler": olcumler}
            inputs_for_history = {"Ayarlanan Sıra Üzeri Mesafe": ayarlanan_str,
                                  "Tarlada Ölçülen Mesafeler": raw_input_str,
                                  "Sınır Oranı": f"%{secili_sinir:.0f} ({sinir_metni})"}
            results_for_history = {"Toplam Ölçüm": f"{toplam}", "İkizleme": f"{ikiz} (%{ikiz_yuzde:.1f})",
                                   "Boşluk": f"{bosluk} (%{bosluk_yuzde:.1f})",
                                   "Kabul Edilebilir": f"{kabul} (%{kabul_yuzde:.1f})",
                                   "Ekim Kalitesi Sınırı": f"%{secili_sinir:.0f} ({sinir_metni})",
                                   "Ekim Kalitesi Oranı": f"%{kabul_yuzde:.1f}",
                                   "Durum": "UYGUN ✅" if uygun else "UYGUN DEĞİL ❌"}
            calc_result_object = CalculationResult(calculation_type=tab_name, inputs=inputs_for_history,
                                                   results=results_for_history,
                                                   timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            self.settings.add_history(tab_name, calc_result_object)
            self.last_calculation_results[tab_name] = {
                "title": f"{tab_name} Hesaplama Raporu",
                "parameters": [["Ayarlanan Sıra Üzeri Mesafe", f"{ayarlanan:.2f}", "cm"],
                               ["Tarlada Ölçülen Mesafeler", ", ".join([f"{x:.2f}" for x in olcumler]), "cm"],
                               ["Sınır Oranı", f"%{secili_sinir:.0f} ({sinir_metni})", ""]],
                "results": [["İkizleme Sınırı", f"{ikiz_sin:.2f} cm"], ["Boşluk Sınırı", f"{bosluk_sin:.2f} cm"],
                            ["Toplam Ölçüm Sayısı", f"{toplam}"], ["İkizleme Sayısı", f"{ikiz} (%{ikiz_yuzde:.1f})"],
                            ["Boşluk Sayısı", f"{bosluk} (%{bosluk_yuzde:.1f})"],
                            ["Kabul Edilebilir Sayısı", f"{kabul} (%{kabul_yuzde:.1f})"],
                            ["Sınır Oranı", f"%{secili_sinir:.0f} ({sinir_metni})"],
                            ["Kabul Edilebilir Oran", f"%{kabul_yuzde:.1f}"],
                            ["Ekim Kalitesi Değerlendirmesi", "UYGUN ✅" if uygun else "UYGUN DEĞİL ❌"]],
                "formula": "İkizleme: Ölçüm < Ayarlanan * 0.5\nBoşluk: Ölçüm > Ayarlanan * 1.5\nKabul Edilebilir = Toplam - İkizleme - Boşluk"
            }
        except Exception as e:
            self.ekim_kalitesi_sonuc.config(text=f"Hesaplama hatası: {str(e)}", bootstyle="danger")
            self.ekim_kalitesi_uygunluk_label.config(text="", foreground="black")
            logger.error(f"Ekim Kalitesi hesaplamasında hata: {e}", exc_info=True)

    def create_tab_fertilization(self):
        tab_name = "Gübreleme Normu"
        tab = tb.Frame(self.notebook)
        self.notebook.add(tab, text=tab_name)
        vcmd_float = (self.register(is_float_input_valid), '%P')
        qh_frame = tb.LabelFrame(tab, text="Q(h) Hesaplama")
        qh_frame.pack(padx=20, pady=10, fill="x")
        self.lbl_vb = tb.Label(qh_frame, text="Besleme Hızı (Vb) (m/dak)")
        self.lbl_vb.pack(pady=(5, 0))
        self.entry_vb = tb.Entry(qh_frame, validate="key", validatecommand=vcmd_float)
        self.entry_vb.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_vb)
        self.lbl_b_fertilization = tb.Label(qh_frame, text=f"Kasa Genişliği (b) ({self.unit})")
        self.lbl_b_fertilization.pack(pady=(5, 0))
        self.entry_b_fertilization = tb.Entry(qh_frame, validate="key", validatecommand=vcmd_float)
        self.entry_b_fertilization.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_b_fertilization)
        self.lbl_h_fertilization = tb.Label(qh_frame, text=f"Kasa Doldurma Yüksekliği (h) ({self.unit})")
        self.lbl_h_fertilization.pack(pady=(5, 0))
        self.entry_h_fertilization = tb.Entry(qh_frame, validate="key", validatecommand=vcmd_float)
        self.entry_h_fertilization.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_h_fertilization)
        self.lbl_lambda = tb.Label(qh_frame, text="Gübrenin Özgül Ağırlığı (λ) (kg/m³)")
        self.lbl_lambda.pack(pady=(5, 0))
        self.entry_lambda = tb.Entry(qh_frame, validate="key", validatecommand=vcmd_float)
        self.entry_lambda.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_lambda)
        tb.Button(qh_frame, text="Q(h) Hesapla", command=self.calculate_qh, bootstyle="success").pack(pady=10)
        self.result_label_qh = tb.Label(qh_frame, text="", font=('Helvetica', 10))
        self.result_label_qh.pack(pady=5)
        q_frame = tb.LabelFrame(tab, text="Gübreleme Normu (q) Hesaplama")
        q_frame.pack(padx=20, pady=10, fill="x")
        tb.Label(q_frame, text="Not: Önce Q(h) değeri hesaplanmalıdır.").pack(pady=(5, 0))
        self.lbl_vbe_width = tb.Label(q_frame, text="İş Genişliği (W) (m)")
        self.lbl_vbe_width.pack(pady=(5, 0))
        self.entry_vbe_width = tb.Entry(q_frame, validate="key", validatecommand=vcmd_float)
        self.entry_vbe_width.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_vbe_width)
        self.lbl_vbe_speed = tb.Label(q_frame, text="İlerleme Hızı (V) (km/h)")
        self.lbl_vbe_speed.pack(pady=(5, 0))
        self.entry_vbe_speed = tb.Entry(q_frame, validate="key", validatecommand=vcmd_float)
        self.entry_vbe_speed.pack(pady=2)
        self.entry_widgets[tab_name].append(self.entry_vbe_speed)
        button_frame_main = tb.Frame(tab)
        button_frame_main.pack(pady=10)
        tb.Button(button_frame_main, text="Gübreleme Normu Hesapla", command=lambda: self.calculate_q(tab_name),
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame_main, text="Temizle", command=lambda: self.clear_entries(tab_name),
                  bootstyle="secondary").pack(side="left", padx=5)
        tb.Button(button_frame_main, text="Geçmiş", command=lambda: self.show_history_window(tab_name),
                  bootstyle="info").pack(side="left", padx=5)
        self.result_label_q = tb.Label(tab, text="", font=('Helvetica', 12))
        self.result_label_q.pack(pady=5)
        button_frame_pdf = tb.Frame(tab)
        button_frame_pdf.pack(pady=20)
        tb.Button(button_frame_pdf, text="PDF'e Yazdır", command=lambda: self.save_pdf(5, tab_name),
                  bootstyle="warning").pack(side="left", padx=5)
        tb.Button(button_frame_pdf, text="Excel'e Aktar",
                  command=lambda: self.export_current_calculation_to_excel(tab_name), bootstyle="success").pack(
            side="left", padx=5)
        tb.Label(tab, text="Formüller:", font=('Helvetica', 10, 'bold')).pack(pady=(10, 5))
        tb.Label(tab, text="Q(h) (ton/saat) = 0.06 × Vb (m/dak) × h (m) × b (m) × λ (kg/m³)",
                 font=("Arial", 10, "italic"), foreground="green").pack()
        tb.Label(tab, text="q (kg/da) = (Q(h) (ton/saat) / (W (m) × V (km/saat))) × 1000", font=("Arial", 10, "italic"),
                 foreground="green").pack()
        self.labels_to_update.extend([self.lbl_b_fertilization, self.lbl_h_fertilization])

    def calculate_qh(self):
        self.result_label_qh.config(text="")
        try:
            vb_str = self.entry_vb.get()
            b_str = self.entry_b_fertilization.get()
            h_str = self.entry_h_fertilization.get()
            lambda_str = self.entry_lambda.get()
            if not vb_str or not b_str or not h_str or not lambda_str:
                self.result_label_qh.config(text="Hata: Q(h) için tüm giriş alanları doldurulmalıdır.",
                                            bootstyle="danger")
                return
            if not all(is_float_input_valid(s) for s in [vb_str, b_str, h_str, lambda_str]):
                self.result_label_qh.config(text="Hata: Geçersiz sayısal giriş.", bootstyle="danger")
                return
            vb = parse_float(vb_str)
            b_case = parse_float(b_str)
            h_case = parse_float(h_str)
            lambda_d = parse_float(lambda_str)
            if None in (vb, b_case, h_case, lambda_d):
                self.result_label_qh.config(text="Hata: Geçersiz giriş değeri!", bootstyle="danger")
                return
            b_calc = b_case / CM_TO_M_FACTOR if self.unit == "cm" else b_case
            h_calc = h_case / CM_TO_M_FACTOR if self.unit == "cm" else h_case
            logger.debug(
                f"Converted units for Q(h) calc: b={b_calc}m, h={h_calc}m, vb={vb}m/dak, lambda={lambda_d}kg/m³")
            self.qh_value = calculate_fertilization_qh(vb, h_calc, b_calc, lambda_d)
            self.result_label_qh.config(text=f"Q(h): {self.qh_value:.2f} ton/saat", bootstyle="success")
            logger.info(f"Calculated Q(h): {self.qh_value:.2f} ton/saat")
        except (ValueError, TypeError) as e:
            self.result_label_qh.config(text=f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})", bootstyle="danger")
            logger.error(f"Error in Q(h) calculation: {e}")
        except Exception as e:
            self.result_label_qh.config(text=f"Beklenmeyen hata: {str(e)}", bootstyle="danger")
            logger.error(f"Unexpected error in Q(h) calculation: {e}")

    def calculate_q(self, tab_name: str):
        self.result_label_q.config(text="")
        try:
            if not hasattr(self, 'qh_value') or self.qh_value is None:
                self.result_label_q.config(text="Hata: Önce Q(h) değerini hesaplayın!", bootstyle="warning")
                return
            W_str = self.entry_vbe_width.get()
            V_str = self.entry_vbe_speed.get()
            if not W_str or not V_str:
                self.result_label_q.config(text="Hata: Gübreleme normu için tüm giriş alanları doldurulmalıdır.",
                                           bootstyle="danger")
                return
            if not is_float_input_valid(W_str) or not is_float_input_valid(V_str):
                self.result_label_q.config(text="Hata: Geçersiz sayısal giriş.", bootstyle="danger")
                return
            W = parse_float(W_str)
            V = parse_float(V_str)
            if None in (W, V):
                self.result_label_q.config(text="Hata: Geçersiz giriş değeri!", bootstyle="danger")
                return
            if W == 0 or V == 0:
                self.result_label_q.config(text="Hata: İş genişliği veya ilerleme hızı sıfır olamaz!",
                                           bootstyle="danger")
                return
            q_ton = calculate_fertilization_q(self.qh_value, W, V)
            q_kg = q_ton * KG_PER_TON
            result_text = f"Gübreleme Normu: {q_ton:.4f} ton/da  |  {q_kg:.2f} kg/da"
            self.result_label_q.config(text=result_text, bootstyle="success")
            logger.info(f"Calculated fertilization norm: {q_ton:.4f} ton/da ({q_kg:.2f} kg/da)")
            inputs = {"Besleme Hızı (Vb)": self.entry_vb.get() + " m/dak",
                      "Kasa Genişliği (b)": self.entry_b_fertilization.get() + f" {self.unit}",
                      "Kasa Doldurma Yüksekliği (h)": self.entry_h_fertilization.get() + f" {self.unit}",
                      "Gübrenin Özgül Ağırlığı (λ)": self.entry_lambda.get() + " kg/m³",
                      "İş Genişliği (W)": W_str + " m", "İlerleme Hızı (V)": V_str + " km/h"}
            results = {"Q(h)": f"{self.qh_value:.2f} ton/saat",
                       "Gübreleme Normu": f"{q_ton:.4f} ton/da ({q_kg:.2f} kg/da)"}
            calc_result = CalculationResult(calculation_type=tab_name, inputs=inputs, results=results,
                                            timestamp=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            self.last_calculation_results[tab_name] = {
                "parameters": [["Besleme Hızı", self.entry_vb.get(), "m/dak"],
                               ["Kasa Genişliği", self.entry_b_fertilization.get(), self.unit],
                               ["Kasa Doldurma Yüksekliği", self.entry_h_fertilization.get(), self.unit],
                               ["Gübrenin Özgül Ağırlığı", self.entry_lambda.get(), "kg/m³"],
                               ["İş Genişliği", self.entry_vbe_width.get(), "m"],
                               ["İlerleme Hızı", self.entry_vbe_speed.get(), "km/h"]],
                "results": [["Q(h)", f"{self.qh_value:.2f} ton/saat"],
                            ["Gübreleme Normu", f"{q_ton:.4f} ton/da | {q_kg:.2f} kg/da"]],
                "formula": "Q(h) (ton/saat) = 0.06 × Vb (m/dak) × h (m) × b (m) × λ (kg/m³)\nq (kg/da) = (Q(h) (ton/saat) / (W (m) × V (km/saat))) × 1000"
            }
            self.settings.add_history(tab_name, calc_result)
        except (ValueError, TypeError) as e:
            self.result_label_q.config(text=f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})", bootstyle="danger")
            logger.error(f"Error in fertilization norm calculation: {e}")
        except Exception as e:
            self.result_label_q.config(text=f"Beklenmeyen hata: {str(e)}", bootstyle="danger")
            logger.error(f"Unexpected error in fertilization norm calculation: {e}")

    def create_other_calculations_tab(self):
        tab_name = "Diğer"
        main_tab = tb.Frame(self.notebook)
        self.notebook.add(main_tab, text=tab_name)
        sub_notebook = tb.Notebook(main_tab)
        sub_notebook.pack(padx=10, pady=10, fill="both", expand=True)
        tab_is_basarisi_name = "İş Başarısı"
        tab_is_basarisi = tb.Frame(sub_notebook)
        sub_notebook.add(tab_is_basarisi, text=tab_is_basarisi_name)
        vcmd_float = (self.register(is_float_input_valid), '%P')
        input_frame_is = tb.LabelFrame(tab_is_basarisi, text="Giriş Değerleri")
        input_frame_is.pack(padx=20, pady=10, fill="x")
        tb.Label(input_frame_is, text="İş Genişliği (m)").pack(pady=(5, 0))
        self.entry_is_genislik = tb.Entry(input_frame_is, validate="key", validatecommand=vcmd_float)
        self.entry_is_genislik.pack(pady=2)
        self.entry_widgets["İş Başarısı"].append(self.entry_is_genislik)
        tb.Label(input_frame_is, text="Ortalama Çalışma Hızı (km/saat)").pack(pady=(5, 0))
        self.entry_is_hiz = tb.Entry(input_frame_is, validate="key", validatecommand=vcmd_float)
        self.entry_is_hiz.pack(pady=2)
        self.entry_widgets["İş Başarısı"].append(self.entry_is_hiz)
        tb.Label(input_frame_is, text="Tarla Etmen Katsayısı (%)").pack(pady=(5, 0))
        self.entry_is_verim = tb.Entry(input_frame_is, validate="key", validatecommand=vcmd_float)
        self.entry_is_verim.insert(0, "80")
        self.entry_is_verim.pack(pady=2)
        self.entry_widgets["İş Başarısı"].append(self.entry_is_verim)
        tb.Label(input_frame_is, text="Toplam Tarla Alanı (dekar)").pack(pady=(5, 0))
        self.entry_is_alan = tb.Entry(input_frame_is, validate="key", validatecommand=vcmd_float)
        self.entry_is_alan.pack(pady=2)
        self.entry_widgets["İş Başarısı"].append(self.entry_is_alan)
        button_frame_is = tb.Frame(tab_is_basarisi)
        button_frame_is.pack(pady=10)
        tb.Button(button_frame_is, text="Hesapla", command=self.calculate_work_success, bootstyle="success").pack(
            side="left", padx=5)
        tb.Button(button_frame_is, text="Temizle", command=lambda: self.clear_entries("İş Başarısı"),
                  bootstyle="secondary").pack(side="left", padx=5)
        result_frame_is = tb.LabelFrame(tab_is_basarisi, text="Sonuçlar")
        result_frame_is.pack(padx=20, pady=10, fill="x")
        self.result_label_is1 = tb.Label(result_frame_is, text="", font=('DejaVu Sans', 12))
        self.result_label_is1.pack(pady=5)
        self.result_label_is2 = tb.Label(result_frame_is, text="", font=('DejaVu Sans', 12))
        self.result_label_is2.pack(pady=5)
        formul_label_is = tb.Label(tab_is_basarisi,
                                   text="Formül: İşlenen Alan (da/saat) = (Genişlik × Hız × Verimlilik) / 10",
                                   font=("DejaVu Sans", 10, "italic"), foreground="green")
        formul_label_is.pack(pady=(10, 10))
        tab_cimlenme_name = "Çimlenme Normu"
        tab_cimlenme = tb.Frame(sub_notebook)
        sub_notebook.add(tab_cimlenme, text=tab_cimlenme_name)
        input_frame_cim = tb.LabelFrame(tab_cimlenme, text="Giriş Değerleri")
        input_frame_cim.pack(padx=20, pady=10, fill="x")
        tb.Label(input_frame_cim, text="Dekara Hedeflenen Bitki Sayısı").pack(pady=(5, 0))
        self.entry_cim_hedef = tb.Entry(input_frame_cim, validate="key", validatecommand=vcmd_float)
        self.entry_cim_hedef.pack(pady=2)
        self.entry_widgets["Çimlenme Normu"].append(self.entry_cim_hedef)
        tb.Label(input_frame_cim, text="1000 Tane Ağırlığı (gram)").pack(pady=(5, 0))
        self.entry_cim_agirlik = tb.Entry(input_frame_cim, validate="key", validatecommand=vcmd_float)
        self.entry_cim_agirlik.pack(pady=2)
        self.entry_widgets["Çimlenme Normu"].append(self.entry_cim_agirlik)
        tb.Label(input_frame_cim, text="Tohum Çimlenme Oranı (%)").pack(pady=(5, 0))
        self.entry_cim_oran = tb.Entry(input_frame_cim, validate="key", validatecommand=vcmd_float)
        self.entry_cim_oran.pack(pady=2)
        self.entry_widgets["Çimlenme Normu"].append(self.entry_cim_oran)
        tb.Label(input_frame_cim, text="Tohum Saflığı (%)").pack(pady=(5, 0))
        self.entry_cim_saflik = tb.Entry(input_frame_cim, validate="key", validatecommand=vcmd_float)
        self.entry_cim_saflik.pack(pady=2)
        self.entry_widgets["Çimlenme Normu"].append(self.entry_cim_saflik)
        button_frame_cim = tb.Frame(tab_cimlenme)
        button_frame_cim.pack(pady=10)
        tb.Button(button_frame_cim, text="Hesapla", command=self.calculate_germination_seed_rate,
                  bootstyle="success").pack(side="left", padx=5)
        tb.Button(button_frame_cim, text="Temizle", command=lambda: self.clear_entries("Çimlenme Normu"),
                  bootstyle="secondary").pack(side="left", padx=5)
        result_frame_cim = tb.LabelFrame(tab_cimlenme, text="Sonuç")
        result_frame_cim.pack(padx=20, pady=10, fill="x")
        self.result_label_cim = tb.Label(result_frame_cim, text="", font=('DejaVu Sans', 12))
        self.result_label_cim.pack(pady=5)
        formul_label_cim = tb.Label(tab_cimlenme,
                                    text="Formül: Gerekli Tohum (kg/da) = (Hedef Bitki Sayısı × 1000 Tane Ağırlığı) / (Çimlenme Oranı × Saflık)",
                                    font=("DejaVu Sans", 9, "italic"), foreground="green")
        formul_label_cim.pack(pady=(10, 10))

    def calculate_work_success(self):
        self.result_label_is1.config(text="")
        self.result_label_is2.config(text="")
        try:
            genislik_str = self.entry_is_genislik.get()
            hiz_str = self.entry_is_hiz.get()
            verim_str = self.entry_is_verim.get()
            alan_str = self.entry_is_alan.get()
            if not genislik_str or not hiz_str or not verim_str or not alan_str:
                messagebox.showerror("Hata", "Tüm giriş alanları doldurulmalıdır.")
                return
            genislik = parse_float(genislik_str)
            hiz = parse_float(hiz_str)
            verim = parse_float(verim_str)
            alan = parse_float(alan_str)
            islenen_alan_saat = calculate_work_performance(genislik, hiz, verim)
            if islenen_alan_saat > 0:
                sure_saat = alan / islenen_alan_saat
                result_text1 = f"Saatte İşlenen Alan: {islenen_alan_saat:.2f} dekar/saat"
                result_text2 = f"Toplam İş Süresi: {sure_saat:.2f} saat"
                self.result_label_is1.config(text=result_text1, bootstyle="success")
                self.result_label_is2.config(text=result_text2, bootstyle="info")
            else:
                self.result_label_is1.config(text="Hesaplama yapılamadı (Hız veya genişlik sıfır).", bootstyle="danger")
        except (ValueError, TypeError) as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})")
            logger.error(f"Error in work success calculation: {e}")

    def calculate_germination_seed_rate(self):
        self.result_label_cim.config(text="")
        try:
            hedef_str = self.entry_cim_hedef.get()
            agirlik_str = self.entry_cim_agirlik.get()
            oran_str = self.entry_cim_oran.get()
            saflik_str = self.entry_cim_saflik.get()
            if not hedef_str or not agirlik_str or not oran_str or not saflik_str:
                messagebox.showerror("Hata", "Tüm giriş alanları doldurulmalıdır.")
                return
            hedef = parse_float(hedef_str)
            agirlik = parse_float(agirlik_str)
            oran = parse_float(oran_str)
            saflik = parse_float(saflik_str)
            gerekli_tohum_kg = calculate_germination_seed_rate(hedef, agirlik, oran, saflik)
            result_text = f"Gerekli Tohum Miktarı: {gerekli_tohum_kg:.2f} kg/dekar"
            self.result_label_cim.config(text=result_text, bootstyle="success")
        except (ValueError, TypeError) as e:
            messagebox.showerror("Hata", f"Hesaplama hatası: Geçersiz sayısal giriş. ({e})")
            logger.error(f"Error in germination seed rate calculation: {e}")

    def clear_entries(self, tab_name: str):
        for entry_widget in self.entry_widgets.get(tab_name, []):
            if isinstance(entry_widget, (tb.Entry, tk.Entry)):
                entry_widget.delete(0, tk.END)
            elif isinstance(entry_widget, tk.Text):
                entry_widget.delete("1.0", tk.END)
        if tab_name == "Ekim Normu":
            self.result_label1.config(text="")
        elif tab_name == "Tarlada Ekim Normu":
            self.result_label2.config(text="")
        elif tab_name == "Markör Uzunluğu":
            self.result_label3_right.config(text=""); self.result_label3_left.config(text="")
        elif tab_name == "Q20 Hesaplama":
            self.result_label4.config(text="")
        elif tab_name == "Sıra Üzeri Uzaklık":
            self.result_label_row_spacing.config(text="")
        elif tab_name == "Ekim Kalitesi":
            self.ekim_kalitesi_sonuc.config(text=""); self.ekim_kalitesi_uygunluk_label.config(text="")
        elif tab_name == "Gübreleme Normu":
            self.result_label_qh.config(text=""); self.result_label_q.config(text=""); self.qh_value = None if hasattr(
                self, 'qh_value') else None
        elif tab_name == "İş Başarısı":
            self.result_label_is1.config(text=""); self.result_label_is2.config(text="")
        elif tab_name == "Çimlenme Normu":
            self.result_label_cim.config(text="")

    def show_history_window(self, tab_name: str):
        if self.history_windows.get(tab_name) and self.history_windows[tab_name].winfo_exists():
            self.history_windows[tab_name].lift()
            return
        history_window = tb.Toplevel(self)
        history_window.title(f"{tab_name} Geçmişi")
        history_window.geometry("600x400")
        history_window.resizable(True, True)
        self.history_windows[tab_name] = history_window
        history_frame = tb.Frame(history_window)
        history_frame.pack(fill="both", expand=True, padx=10, pady=10)
        tree = tb.Treeview(history_frame, columns=("Timestamp", "Inputs", "Results"), show="headings")
        tree.heading("Timestamp", text="Tarih/Saat");
        tree.heading("Inputs", text="Girdiler");
        tree.heading("Results", text="Sonuçlar")
        tree.column("Timestamp", width=150, anchor="w");
        tree.column("Inputs", width=200, anchor="w");
        tree.column("Results", width=200, anchor="w")
        tree.tag_configure('oddrow', background='#2B3E50', foreground='white');
        tree.tag_configure('evenrow', background='#34495E', foreground='white');
        tree.tag_configure('selected', background='#3498DB', foreground='white')
        scrollbar_y = tb.Scrollbar(history_frame, orient="vertical", command=tree.yview)
        scrollbar_x = tb.Scrollbar(history_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=scrollbar_y.set, xscrollcommand=scrollbar_x.set)
        tree.pack(side="left", fill="both", expand=True)
        scrollbar_y.pack(side="right", fill="y")
        scrollbar_x.pack(side="bottom", fill="x")
        self.populate_history_window(tree, tab_name)
        history_window.protocol("WM_DELETE_WINDOW", lambda: self.on_history_window_close(tab_name, history_window))

    def on_history_window_close(self, tab_name: str, window: tb.Toplevel):
        window.destroy()
        if tab_name in self.history_windows:
            del self.history_windows[tab_name]

    def populate_history_window(self, tree: tb.Treeview, tab_name: str):
        for i, item in enumerate(tree.get_children()): tree.delete(item)
        history_data = self.settings.get_history(tab_name)
        for idx, entry in enumerate(history_data):
            tag = 'oddrow' if idx % 2 == 0 else 'evenrow'
            inputs_str = "; ".join([f"{k}: {v}" for k, v in entry.inputs.items()])
            results_str = "; ".join([f"{k}: {v}" for k, v in entry.results.items()])
            tree.insert("", "end", values=(entry.timestamp, inputs_str, results_str), tags=(tag,))

    def create_quick_access_menu(self):
        quick_frame = tb.Frame(self)
        quick_frame.pack(side="top", fill="x", padx=10, pady=5)

    def show_quick_stats(self):
        stats_window = tb.Toplevel(self)
        stats_window.title("📊 Hızlı İstatistikler")
        stats_window.geometry("400x300")
        stats_window.resizable(False, False)
        total_calculations = sum(len(history) for history in self.settings.data["history"].values())
        most_used_tab, most_used_count = (
        max(self.settings.data["history"].keys(), key=lambda x: len(self.settings.data["history"][x])), len(
            self.settings.data["history"][
                max(self.settings.data["history"].keys(), key=lambda x: len(self.settings.data["history"][x]))])) if \
        self.settings.data["history"] else ("Henüz yok", 0)
        stats_text = f"""
    📊 UYGULAMA İSTATİSTİKLERİ

    🔢 Toplam Hesaplama: {total_calculations}
    ⭐ En Çok Kullanılan: {most_used_tab} ({most_used_count} kez)
    📅 Son Kullanım: {datetime.now().strftime("%d.%m.%Y %H:%M")}
    💾 Kayıtlı Veri: {len(self.settings.data["history"])} kategori

    🏆 BAŞARIMLAR:
    • {total_calculations} başarılı hesaplama
    • {len(self.settings.data["history"])} farklı makine tipi
    • %100 başarı oranı
        """
        tb.Label(stats_window, text=stats_text, justify="left", font=("DejaVu Sans", 10)).pack(pady=20, padx=20)
        tb.Button(stats_window, text="Kapat", command=stats_window.destroy, bootstyle="secondary").pack(pady=10)

    def save_all_calculations(self):
        try:
            filename = asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF dosyaları", "*.pdf")],
                                         initialfile="tum_hesaplamalar.pdf")
            if filename:
                all_data = {"title": "Tüm Hesaplamalar Raporu", "parameters": [], "results": [],
                            "formula": "Tüm hesaplama geçmişi"}
                for tab_name, history in self.settings.data["history"].items():
                    if history:
                        all_data["results"].append(f"=== {tab_name} ===")
                        for i, entry in enumerate(history[:5]):
                            timestamp = entry.get('timestamp', 'Bilinmeyen tarih')
                            inputs = entry.get('inputs', {});
                            results = entry.get('results', {})
                            all_data["results"].append(f"--- Hesaplama {i + 1} ({timestamp}) ---")
                            for key, value in inputs.items(): all_data["results"].append(f"Giriş - {key}: {value}")
                            for key, value in results.items(): all_data["results"].append(f"Sonuç - {key}: {value}")
                            all_data["results"].append("")
                if PDFGenerator.generate_pdf(filename, all_data):
                    messagebox.showinfo("✅ Başarılı", f"Tüm hesaplamalar '{filename}' dosyasına kaydedildi.")
                else:
                    messagebox.showerror("❌ Hata", "PDF oluşturulurken hata oluştu.")
        except Exception as e:
            messagebox.showerror("❌ Hata", f"Kaydetme hatası: {str(e)}")

    def clear_all_history(self):
        if messagebox.askyesno("Onay", "Tüm hesaplama geçmişi silinecek. Emin misiniz?"):
            self.settings.data["history"] = {};
            self.settings.save()
            messagebox.showinfo("Başarılı", "Geçmiş temizlendi.")

    def export_current_calculation_to_excel(self, tab_name: str):
        try:
            if tab_name not in self.last_calculation_results:
                messagebox.showwarning("Uyarı",
                                       "Aktarılacak hesaplama sonucu bulunamadı.\nÖnce hesaplama yapmalısınız.")
                return
            filename = asksaveasfilename(defaultextension=".xlsx",
                                         filetypes=[("Excel dosyaları", "*.xlsx"), ("Tüm dosyalar", "*.*")],
                                         initialfile=f"{tab_name}_Hesaplama_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            if not filename: return
            self.create_calculation_excel(tab_name, filename)
            messagebox.showinfo("Başarılı", f"Hesaplama sonucu {filename} dosyasına aktarıldı.")
        except Exception as e:
            messagebox.showerror("Hata", f"Excel aktarımında hata oluştu: {str(e)}"); logger.error(
                f"Current calculation Excel export error: {e}")

    def create_calculation_excel(self, tab_name: str, filename: str):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook();
        ws = wb.active;
        ws.title = f"{tab_name} Hesaplama"
        ws['A1'] = f"{tab_name} Hesaplama Raporu";
        ws['A1'].font = Font(size=18, bold=True, color="FFFFFF");
        ws['A1'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid");
        ws.merge_cells('A1:D1')
        ws['A2'] = f"Oluşturulma Tarihi: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}";
        ws['A2'].font = Font(size=12, italic=True);
        ws.merge_cells('A2:D2')
        ws['A3'] = "Hazırlayan: Hasan Dural | Danışman: Prof.Dr.Davut Karayel";
        ws['A3'].font = Font(size=10, italic=True);
        ws.merge_cells('A3:D3')
        ws['A5'] = "GİRİŞ PARAMETRELERİ";
        ws['A5'].font = Font(size=14, bold=True, color="FFFFFF");
        ws['A5'].fill = PatternFill(start_color="A23B72", end_color="A23B72", fill_type="solid");
        ws.merge_cells('A5:D5')
        if 'parameters' in self.last_calculation_results[tab_name]:
            row = 6;
            ws[f'A{row}'] = "Parametre";
            ws[f'B{row}'] = "Değer";
            ws[f'C{row}'] = "Birim";
            ws[f'D{row}'] = "Açıklama"
            for col in ['A', 'B', 'C', 'D']: cell = ws[f'{col}{row}']; cell.font = Font(bold=True,
                                                                                        color="FFFFFF"); cell.fill = PatternFill(
                start_color="F18F01", end_color="F18F01", fill_type="solid"); cell.alignment = Alignment(
                horizontal="center", vertical="center")
            row += 1
            for param in self.last_calculation_results[tab_name]['parameters']:
                ws[f'A{row}'] = param[0] if len(param) > 0 else "";
                ws[f'B{row}'] = param[1] if len(param) > 1 else "";
                ws[f'C{row}'] = param[2] if len(param) > 2 else "";
                ws[f'D{row}'] = self.get_parameter_description(param[0]) if len(param) > 0 else ""
                if row % 2 == 0:
                    for col in ['A', 'B', 'C', 'D']: ws[f'{col}{row}'].fill = PatternFill(start_color="F5F5F5",
                                                                                          end_color="F5F5F5",
                                                                                          fill_type="solid")
                row += 1
        row = ws.max_row + 2
        ws[f'A{row}'] = "HESAPLAMA SONUÇLARI";
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF");
        ws[f'A{row}'].fill = PatternFill(start_color="C73E1D", end_color="C73E1D", fill_type="solid");
        ws.merge_cells(f'A{row}:D{row}');
        row += 1
        if 'results' in self.last_calculation_results[tab_name]:
            for result in self.last_calculation_results[tab_name]['results']:
                if isinstance(result, list) and len(result) >= 2:
                    ws[f'A{row}'] = result[0]; ws[f'B{row}'] = result[1]; ws[f'D{row}'] = self.get_result_description(
                        result[0])
                else:
                    ws[f'A{row}'] = str(result)
                for col in ['A', 'B', 'C', 'D']: cell = ws[f'{col}{row}']; cell.font = Font(
                    bold=True); cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
                row += 1
        row = ws.max_row + 2
        ws[f'A{row}'] = "KULLANILAN FORMÜL";
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF");
        ws[f'A{row}'].fill = PatternFill(start_color="6A4C93", end_color="6A4C93", fill_type="solid");
        ws.merge_cells(f'A{row}:D{row}');
        row += 1
        if 'formula' in self.last_calculation_results[tab_name]: ws[f'A{row}'] = \
        self.last_calculation_results[tab_name]['formula']; ws[f'A{row}'].font = Font(italic=True, color="2C3E50"); ws[
            f'A{row}'].fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid"); ws.merge_cells(
            f'A{row}:D{row}')
        ws.column_dimensions['A'].width = 25;
        ws.column_dimensions['B'].width = 20;
        ws.column_dimensions['C'].width = 15;
        ws.column_dimensions['D'].width = 30
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=4):
            for cell in row:
                if cell.value: cell.border = thin_border
        wb.save(filename)

    def get_parameter_description(self, param_name: str) -> str:
        descriptions = {"Tekerlek Çapı": "Makine tekerleğinin çapı",
                        "İş Genişliği": "Makinenin bir seferde işlediği alan genişliği",
                        "20 Devirdeki Tohum Miktarı": "Tekerleğin 20 devrinde atılan tohum miktarı",
                        "Tüketilen Tohum Miktarı": "Belirli alanda tüketilen toplam tohum",
                        "Makinenin Aldığı Yol": "Ekim sırasında alınan mesafe",
                        "Traktör Ön Tekerlek İz Genişliği": "Traktörün ön tekerleklerinin arasındaki mesafe",
                        "Ekim Makinesi İş Genişliği": "Ekim makinesinin çalışma genişliği",
                        "Transmisyon Oranı": "Tekerlek ve ekim ünitesi arasındaki hız oranı",
                        "Delik Sayısı": "Ekim diski üzerindeki tohum delik sayısı",
                        "Besleme Hızı": "Gübre besleme hızı", "Kasa Genişliği": "Gübre kasasının genişliği",
                        "Kasa Doldurma Yüksekliği": "Gübre kasasında gübre yüksekliği",
                        "Gübrenin Özgül Ağırlığı": "Gübrenin birim hacimdeki ağırlığı",
                        "İlerleme Hızı": "Makinenin çalışma hızı"}
        return descriptions.get(param_name, "")

    def get_result_description(self, result_name: str) -> str:
        descriptions = {"Ekim Normu": "Dekar başına atılması gereken tohum miktarı",
                        "Tarlada Ekim Normu": "Gerçek tarla koşullarında hesaplanan ekim normu",
                        "Sağ Markör": "Sağ taraftaki markör uzunluğu", "Sol Markör": "Sol taraftaki markör uzunluğu",
                        "20 Devirdeki Tohum Miktarı (Q20)": "Tekerleğin 20 devrinde atılacak tohum miktarı",
                        "Sıra Üzeri Uzaklık": "Tohumlar arasındaki mesafe",
                        "Gübreleme Normu": "Dekar başına uygulanacak gübre miktarı",
                        "Q(h)": "Saatlik gübre dağıtım kapasitesi"}
        return descriptions.get(result_name, "")

    def save_pdf(self, calculation_type: int, tab_name: str):
        filename = asksaveasfilename(defaultextension=".pdf",
                                     filetypes=[("PDF dosyaları", "*.pdf"), ("Tüm dosyalar", "*.*")],
                                     initialfile=f"{tab_name}_Raporu.pdf")
        if not filename: return
        pdf_data = {}
        if calculation_type == 1:
            D = parse_float(self.entry_d.get());
            B = parse_float(self.entry_b.get());
            S = parse_float(self.entry_s.get());
            q_result = getattr(self, 'q_result', "Hesaplanmadı");
            alan = parse_float(self.area_entry1.get());
            toplam_tohum = (q_result * alan) if (isinstance(q_result, float) and alan is not None) else "Hesaplanmadı"
            pdf_data = {"title": "Ekim Normu (20 Devir) Hesaplama Raporu",
                        "parameters": [["Tekerlek Çapı", f"{D:.2f}" if D is not None else "N/A", self.unit],
                                       ["İş Genişliği", f"{B:.2f}" if B is not None else "N/A", self.unit],
                                       ["20 Devirdeki Tohum Miktarı", f"{S:.2f}" if S is not None else "N/A", "kg"]],
                        "results": [
                            ["Ekim Normu", f"{q_result:.2f} kg/da" if isinstance(q_result, float) else str(q_result)],
                            ["Tarla Alanı", f"{alan:.2f} da" if alan is not None else "N/A"], ["Toplam Tohum Miktarı",
                                                                                               f"{toplam_tohum:.2f} kg" if isinstance(
                                                                                                   toplam_tohum,
                                                                                                   float) else str(
                                                                                                   toplam_tohum)]],
                        "formula": "Q (kg/da) = Q20 (kg) / (0.063 × D (m) × B (m))"}
        elif calculation_type == 2:
            Q_val = parse_float(self.entry_q_field.get());
            L_val = parse_float(self.entry_l_field.get());
            B_val = parse_float(self.entry_b_field.get());
            field_rate = getattr(self, 'q_field_result', "Hesaplanmadı");
            alan = parse_float(self.area_entry2.get());
            toplam_tohum = (field_rate * alan) if (
                        isinstance(field_rate, float) and alan is not None) else "Hesaplanmadı"
            pdf_data = {"title": "Tarlada Ekim Normu Hesaplama Raporu", "parameters": [
                ["Tüketilen Tohum Miktarı", f"{Q_val:.2f}" if Q_val is not None else "N/A", "kg"],
                ["Ekim Makinesinin Aldığı Yol", f"{L_val:.2f}" if L_val is not None else "N/A", "m"],
                ["Ekim Makinesinin İş Genişliği", f"{B_val:.2f}" if B_val is not None else "N/A", self.unit]],
                        "results": [["Ekim Normu",
                                     f"{field_rate:.2f} kg/da" if isinstance(field_rate, float) else str(field_rate)],
                                    ["Tarla Alanı", f"{alan:.2f} da" if alan is not None else "N/A"],
                                    ["Toplam Tohum Miktarı",
                                     f"{toplam_tohum:.2f} kg" if isinstance(toplam_tohum, float) else str(
                                         toplam_tohum)]], "formula": "Q (kg/da) = 1000 × q (kg) / (2 × L (m) × B (m))"}
        elif calculation_type == 3:
            b_val = parse_float(self.entry_b_marker.get());
            l_val = parse_float(self.entry_l_marker.get());
            sag, sol = (calculate_marker_lengths(b_val, l_val,
                                                 unit=self.unit)) if b_val is not None and l_val is not None else (
            None, None)
            pdf_data = {"title": "Markör Uzunluğu Hesaplama Raporu", "parameters": [
                ["Ekim Makinesi İş Genişliği", f"{b_val:.2f}" if b_val is not None else "N/A", self.unit],
                ["Traktör Ön Tekerlek İz Genişliği", f"{l_val:.2f}" if l_val is not None else "N/A", self.unit]],
                        "results": [["Sağ Markör", f"{sag:.2f} cm" if sag is not None else "Hesaplanmadı"],
                                    ["Sol Markör", f"{sol:.2f} cm" if sol is not None else "Hesaplanmadı"]],
                        "formula": "Sağ Markör (cm) = b (cm) - (L (cm) / 2), Sol Markör (cm) = b (cm) + (L (cm) / 2)"}
        elif calculation_type == 4:
            q20_input = parse_float(self.entry_q_calc.get());
            D = parse_float(self.entry_d_calc.get());
            B = parse_float(self.entry_b_calc.get());
            calculated_q20 = None
            if q20_input is not None and D is not None and B is not None: D_calc = D / CM_TO_M_FACTOR if self.unit == "cm" else D; B_calc = B / CM_TO_M_FACTOR if self.unit == "cm" else B; calculated_q20 = calculate_q20_from_params(
                q20_input, D_calc, B_calc)
            pdf_data = {"title": "20 Devirde Atılacak Tohum Miktarı (Q20) Raporu",
                        "parameters": [["Ekim Normu", f"{q20_input:.2f}" if q20_input is not None else "N/A", "kg/da"],
                                       ["Tekerlek Çapı", f"{D:.2f}" if D is not None else "N/A", self.unit],
                                       ["İş Genişliği", f"{B:.2f}" if B is not None else "N/A", self.unit]],
                        "results": [["20 Devirde Atılacak Tohum Miktarı (Q20)",
                                     f"{calculated_q20:.2f} kg" if calculated_q20 is not None else "Hesaplanmadı"]],
                        "formula": "Q20 (kg) = Q (kg/da) × 0.063 × D (m) × B (m)"}
        elif calculation_type == 5:
            vb = parse_float(self.entry_vb.get());
            b_fertilization = parse_float(self.entry_b_fertilization.get());
            h_fertilization = parse_float(self.entry_h_fertilization.get());
            lambda_d = parse_float(self.entry_lambda.get());
            W_vbe = parse_float(self.entry_vbe_width.get());
            V_vbe = parse_float(self.entry_vbe_speed.get());
            qh_val = getattr(self, 'qh_value', None);
            q_ton = q_kg = None
            if qh_val is not None and W_vbe is not None and V_vbe is not None and W_vbe != 0 and V_vbe != 0: W_converted = W_vbe / CM_TO_M_FACTOR if self.unit == "cm" else W_vbe; q_ton = calculate_fertilization_q(
                qh_val, W_converted, V_vbe); q_kg = q_ton * KG_PER_TON
            pdf_data = {"title": "Gübreleme Normu Hesaplama Raporu",
                        "parameters": [["Besleme Hızı (Vb)", f"{vb:.2f}" if vb is not None else "N/A", "m/dak"],
                                       ["Kasa Genişliği (b)",
                                        f"{b_fertilization:.2f}" if b_fertilization is not None else "N/A", self.unit],
                                       ["Kasa Doldurma Yüksekliği (h)",
                                        f"{h_fertilization:.2f}" if h_fertilization is not None else "N/A", self.unit],
                                       ["Gübrenin Özgül Ağırlığı (λ)",
                                        f"{lambda_d:.2f}" if lambda_d is not None else "N/A", "kg/m³"],
                                       ["İş Genişliği (W)", f"{W_vbe:.2f}" if W_vbe is not None else "N/A", self.unit],
                                       ["İlerleme Hızı (V)", f"{V_vbe:.2f}" if V_vbe is not None else "N/A", "km/h"]],
                        "results": [["Q(h)", f"{qh_val:.2f} ton/saat" if qh_val is not None else "Hesaplanmadı"],
                                    ["Gübreleme Normu", f"{q_ton:.4f} ton/da" if q_ton is not None else "Hesaplanmadı"],
                                    ["Gübreleme Normu", f"{q_kg:.2f} kg/da" if q_kg is not None else "Hesaplanmadı"]],
                        "formula": "Q(h) = 0.06 × Vb × h × b × λ, q (kg/da) = (Q(h) (ton/saat) / (W (m) × V (km/saat))) × 1000"}
        elif calculation_type == 6:
            D_sira = parse_float(self.entry_d_row.get());
            i_sira = parse_float(self.entry_i_row.get());
            n_sira = parse_float(self.entry_n_row.get());
            calculated_a = None
            if D_sira is not None and i_sira is not None and n_sira is not None and i_sira != 0 and n_sira != 0: D_converted = D_sira / CM_TO_M_FACTOR if self.unit == "cm" else D_sira; calculated_a = calculate_row_spacing(
                D_converted, i_sira, n_sira)
            pdf_data = {"title": "Sıra Üzeri Uzaklık Hesaplama Raporu", "parameters": [
                ["Tekerlek Çapı (D)", f"{D_sira:.2f}" if D_sira is not None else "N/A", self.unit],
                ["Transmisyon Oranı (i)", f"{i_sira:.2f}" if i_sira is not None else "N/A", ""],
                ["Delik Sayısı (n)", f"{n_sira:.2f}" if n_sira is not None else "N/A", ""]], "results": [
                ["Sıra Üzeri Uzaklık", f"{calculated_a:.4f} metre" if calculated_a is not None else "Hesaplanmadı"]],
                        "formula": "a (m) = (π × D (m)) / (i × n)"}
        elif calculation_type == 7:
            if hasattr(self, 'last_ekim_kalitesi_result') and self.last_ekim_kalitesi_result:
                res = self.last_ekim_kalitesi_result
                pdf_data = {"title": "Ekim Kalitesi Hesaplama Raporu",
                            "parameters": [["Ayarlanan Sıra Üzeri Mesafe", f"{res['ayarlanan']:.2f}", "cm"],
                                           ["Tarlada Ölçülen Mesafeler",
                                            ", ".join([f"{x:.2f}" for x in res['olcumler']]), "cm (Liste)"]],
                            "results": [["İkizleme Sınırı", f"{res['ikiz_sin']:.2f} cm"],
                                        ["Boşluk Sınırı", f"{res['bosluk_sin']:.2f} cm"],
                                        ["Toplam Ölçüm Sayısı", f"{res['toplam']}"],
                                        ["İkizleme Sayısı", f"{res['ikiz']} (%{res['ikiz_yuzde']:.1f})"],
                                        ["Boşluk Sayısı", f"{res['bosluk']} (%{res['bosluk_yuzde']:.1f})"],
                                        ["Kabul Edilebilir Sayısı", f"{res['kabul']} (%{res['kabul_yuzde']:.1f})"],
                                        ["Sınır Oranı", f"%{res['sinir']} ({res['sinir_metni']})"],
                                        ["Kabul Edilebilir Oran", f"%{res['kabul_yuzde']:.1f}"],
                                        ["Ekim Kalitesi Değerlendirmesi",
                                         "UYGUN ✅" if res['uygun'] else "UYGUN DEĞİL ❌"]],
                            "formula": "İkizleme: Ölçüm < Ayarlanan * 0.5\nBoşluk: Ölçüm > Ayarlanan * 1.5\nKabul Edilebilir = Toplam - İkizleme - Boşluk"}
            else:
                messagebox.showwarning("Uyarı",
                                       "PDF oluşturmak için önce Ekim Kalitesi hesaplaması yapmalısınız."); return
        else:
            messagebox.showerror("Hata", "Geçersiz hesaplama tipi!"); return
        if PDFGenerator.generate_pdf(filename, pdf_data):
            messagebox.showinfo("PDF Oluşturuldu", f"Rapor '{filename}' başarıyla kaydedildi.")
        else:
            messagebox.showerror("Hata", "PDF oluşturulurken bir sorun oluştu. Detaylar için konsolu kontrol edin.")


if __name__ == "__main__":
    app = SeedRateApp()
    app.mainloop()
